#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Hotword WER evaluation script.
Computes overall WER, U-WER (unbiased, non-hotword tokens),
and B-WER (biased, hotword tokens), plus hotword recall (TP/TN/FP/FN).

Data format:
  - result_dir: entity-annotated reference JSON files, e.g.
      result_dir/AGR-CH/AGR-CH#bilibili#BV17ae1zZEaU.json
      Each file: {"audio_name": ..., "segments": [{..., "entities": [...], "start": ..., "end": ..., "text": ...}]}
  - hyp_dir: model hypothesis JSON files, e.g.
      hyp_dir/AGR-CH_Gemini.json  or  hyp_dir/AGR-CH.json  (model=azure)
      Each file: list of {"audio_name": "...wav", "text": ..., "start_time": ..., "end_time": ..., "language": ..., "model": ...}

Segment matching: by (audio_name_without_wav, start_time, end_time).
"""

import argparse
import json
import os
import sys
import unicodedata
from collections import defaultdict
from multiprocessing import Pool, cpu_count

import kaldialign

ERR = "*"


# ---------------------------------------------------------------------------
# Tokenisation helpers (from wer.py)
# ---------------------------------------------------------------------------

PUNCTS = set("!,?、。！，；？：「」︰『』《》")
SPACELIST = set(" \t\r\n")


def characterize(string: str):
    """Tokenise a string character-by-character (for CJK / mixed text)."""
    res = []
    i = 0
    while i < len(string):
        char = string[i]
        if char in PUNCTS:
            i += 1
            continue
        cat1 = unicodedata.category(char)
        if cat1 == "Zs" or cat1 == "Cn" or char in SPACELIST:
            i += 1
            continue
        if cat1 == "Lo":  # Letter-Other (CJK etc.)
            res.append(char)
            i += 1
        else:
            sep = ">"  if char == "<" else " "
            j = i + 1
            while j < len(string):
                c = string[j]
                if ord(c) >= 128 or c in SPACELIST or c == sep:
                    break
                j += 1
            if j < len(string) and string[j] == ">":
                j += 1
            res.append(string[i:j])
            i = j
    return res


def tokenize(text: str, tochar: bool):
    """Tokenise text; returns list of upper-cased tokens."""
    if tochar:
        tokens = characterize(text)
    else:
        tokens = text.strip().split()
    return [t.upper() for t in tokens]


def entity_tokens(entities, tochar: bool):
    """Return a flat set of tokens from all entity strings."""
    tokens = set()
    for e in entities:
        tokens.update(tokenize(e, tochar))
    return tokens


# ---------------------------------------------------------------------------
# WER accumulator
# ---------------------------------------------------------------------------

class WordError:
    def __init__(self):
        self.sub = 0
        self.ins = 0
        self.dele = 0
        self.ref_words = 0

    def get_wer(self):
        if self.ref_words == 0:
            return 0.0
        return 100.0 * (self.sub + self.ins + self.dele) / self.ref_words

    def __str__(self):
        return (
            f"WER={self.get_wer():.4f}%, "
            f"ref={self.ref_words}, "
            f"sub={self.sub}, "
            f"ins={self.ins}, "
            f"del={self.dele}"
        )


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_ref(result_dir: str, domain: str, norm_ref_dir: str = None):
    """
    Load entity-annotated reference for a domain.
    Returns: dict {(audio_name, start, end): {"text": str, "entities": list}}

    When norm_ref_dir is given (preferred):
      - Segments and text come from norm_ref_dir/{domain}.json
        (same segments as normal WER evaluation)
      - Entities are looked up from result_dir/{domain}/metadata.json
      This ensures segment sets are identical to normal WER.

    Fallback layouts when norm_ref_dir is not given or file missing:
    1. Release metadata: result_dir/{domain}/metadata.json
    2. Legacy entity_ref: result_dir/{domain}/*.json
    """
    # --- Primary: use norm_ref_dir as segment source, metadata for entities ---
    if norm_ref_dir:
        norm_path = os.path.join(norm_ref_dir, f"{domain}.json")
        if os.path.isfile(norm_path):
            # Build entity index from release metadata
            entity_index = {}  # (aid, begin_time, end_time) -> entities list
            domain_dir = os.path.join(result_dir, domain)
            meta_path = os.path.join(domain_dir, "metadata.json") if os.path.isdir(domain_dir) else ""
            if os.path.isfile(meta_path):
                with open(meta_path, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                for audio in meta.get("audios", []):
                    aid = audio.get("aid", "")
                    for seg in audio.get("segments", []):
                        ek = (aid, seg.get("begin_time"), seg.get("end_time"))
                        entity_index[ek] = seg.get("entities", [])

            # Load segments from normalised ref (same set as normal WER)
            ref = {}
            with open(norm_path, "r", encoding="utf-8") as f:
                for item in json.load(f):
                    key = (item["audio_name"], item["start"], item["end"])
                    ref[key] = {
                        "text": item.get("text", ""),
                        "entities": entity_index.get(key, []),
                    }
            return ref

    # --- Fallback: release metadata only ---
    domain_dir = os.path.join(result_dir, domain)
    if not os.path.isdir(domain_dir):
        return {}

    meta_path = os.path.join(domain_dir, "metadata.json")
    if os.path.isfile(meta_path):
        ref = {}
        with open(meta_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for audio in data.get("audios", []):
            aid = audio.get("aid", "")
            for seg in audio.get("segments", []):
                start = seg.get("begin_time")
                end = seg.get("end_time")
                if start is None or end is None:
                    continue
                key = (aid, start, end)
                ref[key] = {
                    "text": seg.get("text", ""),
                    "entities": seg.get("entities", []),
                }
        return ref

    # --- Fallback: legacy entity_ref layout ---
    ref = {}
    for fname in os.listdir(domain_dir):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(domain_dir, fname)
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        audio_name = data["audio_name"]
        for seg in data.get("segments", []):
            if seg.get("status") != "valid":
                continue
            key = (audio_name, seg["start"], seg["end"])
            ref[key] = {
                "text": seg.get("text", ""),
                "entities": seg.get("entities", []),
            }
    return ref


def extract_audio_name(item: dict) -> str:
    """
    Extract the bare audio name (no extension, no directory) from a hyp item.
    Handles several field naming conventions used by different models:
      - item["audio_name"]: may be "AGR-CH#foo#bar.wav" or a full path
      - item["audio_path"]: full Windows/Linux path
    """
    raw = item.get("audio_name") or item.get("audio_path") or ""
    # strip directory components (handles both '/' and '\\')
    raw = raw.replace("\\", "/")
    raw = raw.split("/")[-1]
    # strip extension
    if "." in raw:
        raw = raw.rsplit(".", 1)[0]
    # strip #raw suffix (matches norm_audio_name in compute_wer_single.py)
    if raw.endswith("#raw"):
        raw = raw[:-4]
    return raw


def load_hyp(hyp_dir: str, domain: str, model: str):
    """
    Load hypothesis for a domain+model combination.
    model='azure' corresponds to file {domain}.json; others: {domain}_{model}.json
    Supports both flat layout (hyp_dir/{domain}_{model}.json) and
    subdirectory layout (hyp_dir/{domain}/{domain}_{model}.json).
    Returns: dict {(audio_name, start, end): str}, or None if file not found.
    """
    fname = f"{domain}_{model}.json"
    # Try subdirectory layout first, then flat
    fpath = os.path.join(hyp_dir, domain, fname)
    if not os.path.isfile(fpath):
        fpath = os.path.join(hyp_dir, fname)
    if fpath is None or not os.path.isfile(fpath):
        return None
    hyp = {}
    with open(fpath, "r", encoding="utf-8") as f:
        data = json.load(f)
    for item in data:
        audio_name = extract_audio_name(item)
        # handle both start_time/end_time and start/end field names
        start = item.get("start_time") if "start_time" in item else item.get("start")
        end = item.get("end_time") if "end_time" in item else item.get("end")
        if start is None or end is None:
            continue
        key = (audio_name, start, end)
        hyp[key] = item.get("text", "")
    return hyp


def detect_language(domain: str) -> bool:
    """Return True (tochar=True) for Chinese domains, False for English."""
    return domain.upper().endswith("-CH")


# ---------------------------------------------------------------------------
# Per-domain evaluation
# ---------------------------------------------------------------------------

def evaluate_domain(ref: dict, hyp: dict, tochar: bool, verbose: bool = False):
    """
    Evaluate one domain's ref vs hyp using kaldialign.
    Returns stats dict.
    """
    wer_acc = WordError()
    u_wer_acc = WordError()
    b_wer_acc = WordError()
    hotword = {"tp": 0, "tn": 0, "fp": 0, "fn": 0}
    n_matched_segs = 0
    n_skipped_segs = 0
    overall = {"all": 0, "cor": 0, "sub": 0, "ins": 0, "del": 0}

    # Build fuzzy index for tolerance matching (0.01s)
    _MATCH_TOL = 0.01
    hyp_by_name = defaultdict(list)
    for (h_name, h_start, h_end), h_text in hyp.items():
        hyp_by_name[h_name].append((float(h_start), float(h_end), h_text))

    def _fuzzy_lookup(key):
        name, start, end = key
        start, end = float(start), float(end)
        if key in hyp:
            return hyp[key]
        for h_start, h_end, h_text in hyp_by_name.get(name, []):
            if abs(h_start - start) <= _MATCH_TOL and abs(h_end - end) <= _MATCH_TOL:
                return h_text
        return None

    for key, ref_info in ref.items():
        hyp_text = _fuzzy_lookup(key)
        if hyp_text is None:
            n_skipped_segs += 1
            continue
        n_matched_segs += 1

        lab = tokenize(ref_info["text"], tochar)
        rec = tokenize(hyp_text, tochar)

        # build hotword token set from entities
        hot_set = entity_tokens(ref_info["entities"], tochar)
        hot_true_list = {t for t in hot_set if t in set(lab)}
        hot_bad_list = hot_set - hot_true_list

        # compute alignment via kaldialign (C++ backend)
        ali = kaldialign.align(lab, rec, ERR)

        # accumulate WER / U-WER / B-WER from alignment
        for ref_tok, hyp_tok in ali:
            if ref_tok == ERR:
                # insertion
                overall["ins"] += 1
                wer_acc.ins += 1
                if hyp_tok in hot_true_list:
                    b_wer_acc.ins += 1
                else:
                    u_wer_acc.ins += 1
            elif hyp_tok == ERR:
                # deletion
                overall["all"] += 1
                overall["del"] += 1
                wer_acc.ref_words += 1
                wer_acc.dele += 1
                if ref_tok in hot_true_list:
                    b_wer_acc.ref_words += 1
                    b_wer_acc.dele += 1
                else:
                    u_wer_acc.ref_words += 1
                    u_wer_acc.dele += 1
            elif ref_tok != hyp_tok:
                # substitution
                overall["all"] += 1
                overall["sub"] += 1
                wer_acc.ref_words += 1
                wer_acc.sub += 1
                if ref_tok in hot_true_list:
                    b_wer_acc.ref_words += 1
                    b_wer_acc.sub += 1
                else:
                    u_wer_acc.ref_words += 1
                    u_wer_acc.sub += 1
            else:
                # correct
                overall["all"] += 1
                overall["cor"] += 1
                wer_acc.ref_words += 1
                if ref_tok in hot_true_list:
                    b_wer_acc.ref_words += 1
                else:
                    u_wer_acc.ref_words += 1

        # accumulate hotword recall stats
        rec_tokens = [t for t in rec]
        for bad_hw in hot_bad_list:
            count = rec_tokens.count(bad_hw)
            if count == 0:
                hotword["tn"] += 1
            else:
                hotword["fp"] += count

        for hw in hot_true_list:
            true_cnt = lab.count(hw)
            rec_cnt = rec_tokens.count(hw)
            if rec_cnt >= true_cnt:
                hotword["tp"] += true_cnt
                hotword["fp"] += rec_cnt - true_cnt
            else:
                hotword["tp"] += rec_cnt
                hotword["fn"] += true_cnt - rec_cnt

    recall = (hotword["tp"] / (hotword["tp"] + hotword["fn"]) * 100
              if hotword["tp"] + hotword["fn"] > 0 else 0.0)

    return {
        "n_matched": n_matched_segs,
        "n_skipped": n_skipped_segs,
        "wer": wer_acc,
        "u_wer": u_wer_acc,
        "b_wer": b_wer_acc,
        "hotword": hotword,
        "recall": recall,
        "overall": overall,
    }


def print_domain_result(domain: str, model: str, stats: dict):
    o = stats["overall"]
    wer_val = (o["ins"] + o["sub"] + o["del"]) * 100.0 / o["all"] if o["all"] else 0.0
    hw = stats["hotword"]
    print(f"\n{'='*70}")
    print(f"Domain: {domain}  Model: {model}")
    print(f"  Segments matched: {stats['n_matched']}  skipped: {stats['n_skipped']}")
    print(f"  WER:   {wer_val:.4f}%  "
          f"N={o['all']} C={o['cor']} S={o['sub']} D={o['del']} I={o['ins']}")
    print(f"  WER:   {stats['wer']}")
    print(f"  U-WER: {stats['u_wer']}")
    print(f"  B-WER: {stats['b_wer']}")
    print(f"  Hotword: tp={hw['tp']} tn={hw['tn']} fp={hw['fp']} fn={hw['fn']}  "
          f"recall={stats['recall']:.2f}%")
    print(f"  >> {stats['wer'].get_wer():.3f}; "
          f"{stats['u_wer'].get_wer():.3f}; "
          f"{stats['b_wer'].get_wer():.3f}; "
          f"{stats['recall']:.2f}%")


# ---------------------------------------------------------------------------
# Aggregate across domains
# ---------------------------------------------------------------------------

def aggregate(stats_list):
    """Sum up stats across multiple domain results."""
    agg_wer = WordError()
    agg_u_wer = WordError()
    agg_b_wer = WordError()
    agg_hw = {"tp": 0, "tn": 0, "fp": 0, "fn": 0}

    for s in stats_list:
        agg_wer.sub += s["wer"].sub
        agg_wer.ins += s["wer"].ins
        agg_wer.dele += s["wer"].dele
        agg_wer.ref_words += s["wer"].ref_words
        agg_u_wer.sub += s["u_wer"].sub
        agg_u_wer.ins += s["u_wer"].ins
        agg_u_wer.dele += s["u_wer"].dele
        agg_u_wer.ref_words += s["u_wer"].ref_words
        agg_b_wer.sub += s["b_wer"].sub
        agg_b_wer.ins += s["b_wer"].ins
        agg_b_wer.dele += s["b_wer"].dele
        agg_b_wer.ref_words += s["b_wer"].ref_words
        for k in agg_hw:
            agg_hw[k] += s["hotword"][k]

    recall = (agg_hw["tp"] / (agg_hw["tp"] + agg_hw["fn"]) * 100
              if agg_hw["tp"] + agg_hw["fn"] > 0 else 0.0)
    return {
        "wer": agg_wer, "u_wer": agg_u_wer, "b_wer": agg_b_wer,
        "hotword": agg_hw, "recall": recall,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def get_args():
    parser = argparse.ArgumentParser(description="Hotword WER evaluation")
    parser.add_argument(
        "--result_dir",
        default=None,
        help="Directory with entity-annotated reference JSONs (one subdir per domain)",
    )
    parser.add_argument(
        "--hyp_dir",
        default=None,
        help="Directory with hypothesis JSON files",
    )
    parser.add_argument(
        "--out_dir",
        default=None,
        help="Output directory for hotword results (default: <base>/data/results_hotword)",
    )
    parser.add_argument(
        "--domain",
        nargs="*",
        default=None,
        help="Domain(s) to evaluate, e.g. AGR-CH AGR-EN. Default: all available.",
    )
    parser.add_argument(
        "--model",
        nargs="*",
        default=None,
        help="Model(s) to evaluate, e.g. azure Gemini. Default: all available.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print per-segment details",
    )
    parser.add_argument(
        "--norm_ref_dir",
        default=None,
        help="Directory with normalised ref JSONs ({domain}.json). "
             "Text from these files overrides raw metadata text.",
    )
    parser.add_argument(
        "--skip_existing",
        action="store_true",
        help="Skip domain+model if hotword_cache.json already exists",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="Number of parallel workers (default: 8)",
    )
    return parser.parse_args()


def discover_domains(result_dir: str, hyp_dir: str):
    """Return sorted list of domains present in both result_dir and hyp_dir."""
    result_domains = set(os.listdir(result_dir)) if os.path.isdir(result_dir) else set()
    hyp_domains = set()
    for entry in os.listdir(hyp_dir):
        # Subdirectory layout: hyp_dir/{domain}/
        if os.path.isdir(os.path.join(hyp_dir, entry)):
            hyp_domains.add(entry)
        # Flat layout: hyp_dir/{domain}_{model}.json
        elif entry.endswith(".json"):
            domain = entry.replace(".json", "").split("_")[0]
            hyp_domains.add(domain)
    return sorted(result_domains & hyp_domains)


def discover_models(hyp_dir: str, domain: str):
    """Return list of models available for a domain."""
    models = set()
    # Check subdirectory layout: hyp_dir/{domain}/{domain}_{model}.json
    subdir = os.path.join(hyp_dir, domain)
    if os.path.isdir(subdir):
        for fname in os.listdir(subdir):
            if not fname.endswith(".json"):
                continue
            base = fname.replace(".json", "")
            if base.startswith(domain + "_"):
                models.add(base[len(domain) + 1:])
    # Also check flat layout: hyp_dir/{domain}_{model}.json
    for fname in os.listdir(hyp_dir):
        if not fname.endswith(".json"):
            continue
        base = fname.replace(".json", "")
        if base.startswith(domain + "_"):
            models.add(base[len(domain) + 1:])
    return sorted(models)


def _run_single_task(task):
    """Execute a single hotword WER task (module-level for pickling)."""
    domain, model, tochar, ref, hyp, m_dir, cache_path = task
    stats = evaluate_domain(ref, hyp, tochar, verbose=False)
    os.makedirs(m_dir, exist_ok=True)
    hw = stats["hotword"]
    with open(os.path.join(m_dir, "hotword_result.txt"), "w", encoding="utf8") as f:
        f.write(f"wer={stats['wer'].get_wer():.4f}\n")
        f.write(f"u_wer={stats['u_wer'].get_wer():.4f}\n")
        f.write(f"b_wer={stats['b_wer'].get_wer():.4f}\n")
        f.write(f"recall={stats['recall']:.4f}\n")
        f.write(f"tp={hw['tp']}\ntn={hw['tn']}\nfp={hw['fp']}\nfn={hw['fn']}\n")
        f.write(f"matched_segments={stats['n_matched']}\n")
        f.write(f"skipped_segments={stats['n_skipped']}\n")
    cache_data = {
        "wer": stats["wer"].get_wer(),
        "u_wer": stats["u_wer"].get_wer(),
        "b_wer": stats["b_wer"].get_wer(),
        "recall": stats["recall"],
        "hotword": hw,
        "matched_segments": stats["n_matched"],
        "skipped_segments": stats["n_skipped"],
    }
    with open(cache_path, "w", encoding="utf8") as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)
    return domain, model, stats, cache_data


def main():
    args = get_args()

    # resolve defaults relative to script location
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    staging_root = os.environ.get("STAGING_ROOT", "")
    result_dir = args.result_dir or (
        os.path.join(staging_root, "Vertical-Domain", "data") if staging_root
        else os.path.join(base_dir, "data", "text", "Vertical-Domain", "ref"))
    hyp_dir = args.hyp_dir or os.path.join(base_dir, "data", "text_normalized", "Vertical-Domain", "hyp")
    norm_ref_dir = args.norm_ref_dir or os.path.join(base_dir, "data", "text_normalized", "Vertical-Domain", "ref")
    out_dir = args.out_dir or os.path.join(base_dir, "data", "results_hotword")
    os.makedirs(out_dir, exist_ok=True)

    # determine domains to evaluate
    if args.domain:
        domains = args.domain
    else:
        domains = discover_domains(result_dir, hyp_dir)
    if not domains:
        print("No matching domains found.", file=sys.stderr)
        sys.exit(1)

    # collect per-model aggregated stats
    model_agg = defaultdict(list)
    # table for Excel: {model: {domain: {"wer":..., "u_wer":..., "b_wer":..., "recall":...}}}
    excel_data = defaultdict(dict)

    # Build task list: [(domain, model, tochar, ref, hyp, m_dir, cache_path)]
    tasks = []
    cached_count = 0
    for domain in domains:
        tochar = detect_language(domain)
        ref = load_ref(result_dir, domain, norm_ref_dir=norm_ref_dir)
        if not ref:
            print(f"[WARN] No ref data for domain {domain}, skipping.", file=sys.stderr)
            continue

        models = args.model if args.model else discover_models(hyp_dir, domain)
        for model in models:
            m_dir = os.path.join(out_dir, domain, model.upper())
            cache_path = os.path.join(m_dir, "hotword_cache.json")

            # Skip if cache exists and --skip_existing
            if args.skip_existing and os.path.exists(cache_path):
                try:
                    cached = json.load(open(cache_path, "r", encoding="utf8"))
                    excel_data[model][domain] = {
                        "wer": cached["wer"],
                        "u_wer": cached["u_wer"],
                        "b_wer": cached["b_wer"],
                        "recall": cached["recall"],
                    }
                    cached_count += 1
                    continue
                except Exception:
                    pass  # cache corrupted, recompute

            hyp = load_hyp(hyp_dir, domain, model)
            if hyp is None:
                continue
            tasks.append((domain, model, tochar, ref, hyp, m_dir, cache_path))

    print(f"Hotword WER: {len(tasks)} tasks to compute, {cached_count} cached (skipped)")

    # Run in parallel using multiprocessing (CPU-bound)
    n_workers = min(args.workers, len(tasks)) if tasks else 1
    if tasks:
        with Pool(n_workers) as pool:
            results = pool.map(_run_single_task, tasks)
        for i, (domain, model, stats, cache_data) in enumerate(results):
            model_agg[model].append(stats)
            excel_data[model][domain] = {
                "wer": cache_data["wer"],
                "u_wer": cache_data["u_wer"],
                "b_wer": cache_data["b_wer"],
                "recall": cache_data["recall"],
            }
            print(f"  [{i+1}/{len(tasks)}] {domain}/{model}: WER={cache_data['wer']:.3f} B-WER={cache_data['b_wer']:.3f} recall={cache_data['recall']:.2f}%")

    # print per-model aggregate
    if len(domains) > 1:
        print(f"\n{'='*70}")
        print("AGGREGATE (all domains)")
        for model, stats_list in sorted(model_agg.items()):
            agg = aggregate(stats_list)
            hw = agg["hotword"]
            print(f"\n  Model: {model}  ({len(stats_list)} domains)")
            print(f"  WER:   {agg['wer']}")
            print(f"  U-WER: {agg['u_wer']}")
            print(f"  B-WER: {agg['b_wer']}")
            print(f"  Hotword: tp={hw['tp']} tn={hw['tn']} fp={hw['fp']} fn={hw['fn']}  "
                  f"recall={agg['recall']:.2f}%")
            print(f"  >> {agg['wer'].get_wer():.3f}; "
                  f"{agg['u_wer'].get_wer():.3f}; "
                  f"{agg['b_wer'].get_wer():.3f}; "
                  f"{agg['recall']:.2f}%")

    # Write Excel with 4 sheets: WER, U-WER, B-WER, Recall
    write_hotword_excel(out_dir, excel_data, domains)


def write_hotword_excel(out_dir, excel_data, domains):
    """Write hotword results to Excel with 4 sheets."""
    try:
        import pandas as pd
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("[WARN] pandas/openpyxl not available, skipping Excel output.")
        return

    metrics = ["wer", "u_wer", "b_wer", "recall"]
    sheet_names = {"wer": "WER", "u_wer": "U-WER", "b_wer": "B-WER", "recall": "Recall"}

    out_xlsx = os.path.join(out_dir, "hotword_results.xlsx")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        for metric in metrics:
            table = {}
            for model, domain_stats in excel_data.items():
                row = {}
                for domain in domains:
                    if domain in domain_stats:
                        row[domain] = round(domain_stats[domain][metric], 2)
                    else:
                        row[domain] = "-"
                table[model] = row
            df = pd.DataFrame.from_dict(table, orient="index")
            if not df.empty:
                df = df.reindex(columns=[d for d in domains if d in df.columns])
                # Add MIN row
                min_vals = {}
                for col in df.columns:
                    numeric = pd.to_numeric(df[col], errors="coerce")
                    if numeric.notna().any():
                        min_vals[col] = numeric.min()
                    else:
                        min_vals[col] = "-"
                df = pd.concat([df, pd.DataFrame(min_vals, index=["MIN"])])
            df.to_excel(writer, sheet_name=sheet_names[metric])

            # Style
            ws = writer.sheets[sheet_names[metric]]
            bold_font = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = center
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = center
            for cell in ws[ws.max_row]:
                cell.font = bold_font
            for col_cells in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = max_len + 3

    print(f"\n✅ Hotword Excel -> {out_xlsx}")


if __name__ == "__main__":
    main()
