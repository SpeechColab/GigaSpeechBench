#!/usr/bin/env python3
"""
Translation Quality Evaluation Pipeline
Evaluates translation quality for Low-Resource-Languages using OpenSTBench TranslationEvaluator.
Metrics: sacreBLEU, chrF++, COMET, BLEURT

Usage:
  python eval_translation.py [--models MODEL1 MODEL2 ...] [--langs LANG1 LANG2 ...]
                             [--directions ENG CHN] [--no_comet] [--no_bleurt]
                             [--out_dir OUTPUT_DIR] [--excel_dir EXCEL_DIR]
"""

import argparse
import json
import logging
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ─────────────── Default paths (overridable via CLI or env vars) ───────────────
_SCRIPT_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
_BASE_DIR = _SCRIPT_DIR.parent.parent

METRIC_KEYS = ["sacreBLEU", "chrF++", "COMET", "BLEURT"]

logger = logging.getLogger("translation_eval")

ALL_LANGS = ["ARE", "DZA", "EGY", "IDN", "IRQ", "MAR", "MYS", "PHL", "SAU", "SYR", "THA", "VNM"]
DIRECTIONS = ["ENG", "CHN"]

# ref field mapping: direction -> metadata field
REF_FIELD = {"ENG": "text_en", "CHN": "text_zh"}
# OpenSTBench target_lang mapping
TARGET_LANG_MAP = {"ENG": "en", "CHN": "zh"}


def load_ref(lang: str, data_root: Path) -> dict:
    """Load ref: returns {(begin_time, end_time): segment_dict, ...}"""
    meta_path = data_root / lang / "metadata.json"
    if not meta_path.exists():
        return {}
    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)
    lookup = {}
    for audio in meta.get("audios", []):
        for seg in audio.get("segments", []):
            bt = round(float(seg.get("begin_time", 0)), 3)
            et = round(float(seg.get("end_time", 0)), 3)
            lookup[(bt, et)] = seg
    return lookup


def load_hyp(model_path: Path, lang: str, direction: str) -> list:
    """Load hyp: returns [{begin_time, end_time, translated_text}, ...]"""
    with open(model_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    key = f"{lang}_to_{direction}"
    return data.get(key, [])


def match_ref_hyp(ref_lookup: dict, hyp_list: list, direction: str):
    """Match ref and hyp, returns (ref_texts, hyp_texts, src_texts, matched_count)"""
    ref_texts = []
    hyp_texts = []
    src_texts = []
    field = REF_FIELD[direction]

    for item in hyp_list:
        bt = round(float(item.get("begin_time", 0)), 3)
        et = round(float(item.get("end_time", 0)), 3)
        ref_seg = ref_lookup.get((bt, et))
        if ref_seg is None:
            continue
        ref_text = ref_seg.get(field, "").strip()
        hyp_text = item.get("translated_text", "").strip()
        if not ref_text or not hyp_text:
            continue
        ref_texts.append(ref_text)
        hyp_texts.append(hyp_text)
        src_texts.append(ref_seg.get("text", ""))

    return ref_texts, hyp_texts, src_texts, len(ref_texts)


def setup_logging(log_dir: Path):
    """Configure logging: output to both file and console"""
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"translation_eval_{ts}.log"

    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    fh = logging.FileHandler(str(log_file), encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    logger.info(f"Log file: {log_file}")
    return log_file


def run_eval(args):
    log_file = setup_logging(Path(args.log_dir))

    from openstbench import TranslationEvaluator

    # Resolve paths
    data_root = Path(args.data_root)
    results_trans = Path(args.results_trans)

    evaluator = TranslationEvaluator(
        use_bleu=True,
        use_chrf=True,
        use_comet=not args.no_comet,
        use_bleurt=not args.no_bleurt,
        device=args.device,
    )

    # Discover all model files
    if args.models:
        model_files = []
        for m in args.models:
            p = results_trans / f"{m}.json"
            if p.exists():
                model_files.append(p)
            else:
                print(f"[WARN] Model file not found: {p}")
    else:
        model_files = sorted(results_trans.glob("*.json"))

    if not model_files:
        logger.error("No model result files found.")
        sys.exit(1)

    langs = args.langs or ALL_LANGS
    directions = args.directions or DIRECTIONS
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Models: {[p.stem for p in model_files]}")
    logger.info(f"Langs: {langs}")
    logger.info(f"Directions: {directions}")
    logger.info(f"COMET: {not args.no_comet}, BLEURT: {not args.no_bleurt}, Device: {args.device}")

    # Determine which metrics are needed
    needed_metrics = {"sacreBLEU", "chrF++"}
    if not args.no_comet:
        needed_metrics.add("COMET")
    if not args.no_bleurt:
        needed_metrics.add("BLEURT")

    # Aggregate results
    all_results = {}

    for model_path in model_files:
        model_name = model_path.stem
        logger.info(f"{'='*60}")
        logger.info(f"Model: {model_name}")
        logger.info(f"{'='*60}")

        # Load existing results
        model_out = out_dir / f"{model_name}.json"
        existing_results = {}
        if model_out.exists():
            with open(model_out, "r", encoding="utf-8") as f:
                existing_results = json.load(f)
            logger.info(f"Loaded existing results from {model_out}")

        model_results = dict(existing_results)

        for direction in directions:
            tgt_lang = TARGET_LANG_MAP[direction]
            logger.info(f"Direction: → {direction}")

            for lang in langs:
                result_key = f"{lang}_to_{direction}"

                # Check if existing results already contain all needed metrics
                existing = existing_results.get(result_key, {})
                existing_scores = existing.get("scores", {})
                missing = needed_metrics - set(existing_scores.keys())
                if not missing and result_key in existing_results:
                    logger.info(f"{lang}: all metrics present, skip")
                    continue

                ref_lookup = load_ref(lang, data_root)
                if not ref_lookup:
                    logger.warning(f"{lang}: no ref data, skip")
                    continue

                hyp_list = load_hyp(model_path, lang, direction)
                if not hyp_list:
                    logger.warning(f"{lang}: no hyp data, skip")
                    continue

                ref_texts, hyp_texts, src_texts, matched = match_ref_hyp(
                    ref_lookup, hyp_list, direction
                )
                if matched == 0:
                    logger.warning(f"{lang}: 0 matched segments, skip")
                    continue

                if missing:
                    logger.info(f"{lang}: {matched} matched, computing missing metrics: {missing}")
                else:
                    logger.info(f"{lang}: {matched} matched segments, evaluating...")

                try:
                    scores = evaluator.evaluate_all(
                        reference=ref_texts,
                        target_text=hyp_texts,
                        source=src_texts if not args.no_comet else None,
                        target_lang=tgt_lang,
                    )
                except Exception as e:
                    logger.error(f"{lang}: evaluation error: {e}")
                    scores = {"error": str(e)}

                # Merge existing scores with new scores
                merged_scores = dict(existing_scores)
                merged_scores.update(scores)

                model_results[result_key] = {
                    "matched_segments": matched,
                    "total_hyp": len(hyp_list),
                    "total_ref": len(ref_lookup),
                    "scores": merged_scores,
                }
                logger.info(f"{lang}: {merged_scores}")

        all_results[model_name] = model_results

        # Save results for each model
        with open(model_out, "w", encoding="utf-8") as f:
            json.dump(model_results, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved: {model_out}")

    # Save aggregated results
    summary_path = out_dir / "summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    logger.info(f"All results saved to {summary_path}")

    # Print summary table
    print_summary_table(all_results, directions, langs)

    # Export to Excel
    excel_dir = Path(args.excel_dir)
    excel_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = excel_dir / f"translation_eval_{ts}.xlsx"
    export_excel(all_results, directions, langs, excel_path)

    logger.info(f"Log file: {log_file}")
    logger.info("Done.")
    evaluator.cleanup()


def print_summary_table(all_results, directions, langs):
    """Print summary comparison table"""
    print(f"\n{'='*100}")
    print("  SUMMARY")
    print(f"{'='*100}")

    models = list(all_results.keys())

    for direction in directions:
        print(f"\n  → {direction}")
        header = f"  {'Lang':<8}"
        for m in models:
            short = m[:16]
            header += f"  {short:>16}" + " " * (len(METRIC_KEYS) * 8 - len(short))
        print(header)

        sub = f"  {'':8}"
        for m in models:
            for mk in METRIC_KEYS:
                sub += f" {mk:>8}"
        print(sub)
        print("  " + "-" * (8 + len(models) * len(METRIC_KEYS) * 9))

        for lang in langs:
            key = f"{lang}_to_{direction}"
            row = f"  {lang:<8}"
            for m in models:
                r = all_results.get(m, {}).get(key, {})
                scores = r.get("scores", {})
                for mk in METRIC_KEYS:
                    v = scores.get(mk, None)
                    if isinstance(v, (int, float)):
                        row += f" {v:>8.2f}"
                    else:
                        row += f" {'N/A':>8}"
            print(row)

        # Average row
        row = f"  {'AVG':<8}"
        for m in models:
            for mk in METRIC_KEYS:
                vals = []
                for lang in langs:
                    key = f"{lang}_to_{direction}"
                    r = all_results.get(m, {}).get(key, {})
                    v = r.get("scores", {}).get(mk, None)
                    if isinstance(v, (int, float)):
                        vals.append(v)
                if vals:
                    row += f" {sum(vals)/len(vals):>8.2f}"
                else:
                    row += f" {'N/A':>8}"
        print("  " + "-" * (8 + len(models) * len(METRIC_KEYS) * 9))
        print(row)


def export_excel(all_results, directions, langs, excel_path):
    """Export results to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    models = list(all_results.keys())

    for di, direction in enumerate(directions):
        ws = wb.active if di == 0 else wb.create_sheet()
        ws.title = f"to_{direction}"

        # Header row 1: model names (merged across metrics)
        col = 2
        for m in models:
            ws.cell(1, col, m).font = Font(bold=True)
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + len(METRIC_KEYS) - 1)
            ws.cell(1, col).alignment = Alignment(horizontal="center")
            col += len(METRIC_KEYS)

        # Header row 2: metric names
        ws.cell(2, 1, "Language").font = Font(bold=True)
        col = 2
        for m in models:
            for mk in METRIC_KEYS:
                ws.cell(2, col, mk).font = Font(bold=True)
                col += 1

        # Data rows
        for ri, lang in enumerate(langs, start=3):
            key = f"{lang}_to_{direction}"
            ws.cell(ri, 1, lang)
            col = 2
            for m in models:
                r = all_results.get(m, {}).get(key, {})
                scores = r.get("scores", {})
                for mk in METRIC_KEYS:
                    v = scores.get(mk, None)
                    if isinstance(v, (int, float)):
                        cell = ws.cell(ri, col, round(v, 2))
                        cell.number_format = '0.00'
                    else:
                        ws.cell(ri, col, "N/A")
                    col += 1

        # Average row
        avg_row = len(langs) + 3
        ws.cell(avg_row, 1, "AVG").font = Font(bold=True)
        col = 2
        for m in models:
            for mk in METRIC_KEYS:
                vals = []
                for lang in langs:
                    key = f"{lang}_to_{direction}"
                    r = all_results.get(m, {}).get(key, {})
                    v = r.get("scores", {}).get(mk, None)
                    if isinstance(v, (int, float)):
                        vals.append(v)
                if vals:
                    cell = ws.cell(avg_row, col, round(sum(vals) / len(vals), 2))
                    cell.number_format = '0.00'
                    cell.font = Font(bold=True)
                else:
                    ws.cell(avg_row, col, "N/A")
                col += 1

        # Auto-width
        for c in ws.columns:
            real_cells = [cell for cell in c if not isinstance(cell, openpyxl.cell.cell.MergedCell)]
            if not real_cells:
                continue
            max_len = max(len(str(cell.value or "")) for cell in real_cells)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 2, 20)

    wb.save(excel_path)
    logger.info(f"Excel saved to {excel_path}")


def main():
    staging_root = os.environ.get("STAGING_ROOT", "")
    default_data_root = os.path.join(staging_root, "Low-Resource-Languages", "data") if staging_root else ""
    default_results_trans = os.path.join(staging_root, "Low-Resource-Languages", "results_trans") if staging_root else ""
    default_out = str(_SCRIPT_DIR / "results")
    default_excel = str(_BASE_DIR / "data" / "excel_output")
    default_log = str(_BASE_DIR / "log")

    parser = argparse.ArgumentParser(description="Translation quality evaluation pipeline")
    parser.add_argument("--data_root", default=default_data_root,
                        help="Root dir with per-lang metadata.json (default: $STAGING_ROOT/Low-Resource-Languages/data)")
    parser.add_argument("--results_trans", default=default_results_trans,
                        help="Dir with model translation JSONs (default: $STAGING_ROOT/Low-Resource-Languages/results_trans)")
    parser.add_argument("--models", nargs="*", default=None,
                        help="Model names (stem of json files in results_trans/)")
    parser.add_argument("--langs", nargs="*", default=None,
                        help="Languages to evaluate (default: all 12)")
    parser.add_argument("--directions", nargs="*", default=None,
                        help="Translation directions: ENG CHN (default: both)")
    parser.add_argument("--no_comet", action="store_true",
                        help="Disable COMET scoring")
    parser.add_argument("--no_bleurt", action="store_true",
                        help="Disable BLEURT scoring")
    parser.add_argument("--device", default="cpu",
                        help="Device for COMET/BLEURT (default: cpu)")
    parser.add_argument("--out_dir", default=default_out,
                        help="Output directory for results")
    parser.add_argument("--excel_dir", default=default_excel,
                        help="Excel output directory")
    parser.add_argument("--log_dir", default=default_log,
                        help="Log output directory")
    args = parser.parse_args()
    run_eval(args)


if __name__ == "__main__":
    main()
