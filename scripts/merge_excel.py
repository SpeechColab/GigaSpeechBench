#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Merge per-module results into a single all_results.xlsx with 6 sheets.

Sheet order (fixed):
    1. Low-Resource-Languages
    2. CH-EN-Dialects
    3. fleurs
    4. common-voice
    5. Vertical-Domain-CH
    6. Vertical-Domain-EN

Cell logic:
    - If matched_segments == ref_segments  -> display WER/CER value
    - If not fully aligned                 -> display "matched/ref" string with red cell background
    - If no data                           -> display "-"
"""

import os
import re
import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# Model whitelist + display name mapping
# =========================
MODEL_ORDER = [
    ("Azure",                   {"AZURE"}),
    ("Chirp 3",                 {"CHIRP3", "CHIRP-3"}),
    ("ElevenLabs Scribe v2",    {"ELEVENLABS_SCRIBE_V2", "ELEVENLABS-SCRIBE-V2"}),
    ("Meta OmniASR 3B",         {"OMNIASR_LLM_3B", "META-OMNIAST-3B", "META-OMNIASR-3B"}),
    ("Qwen3-ASR-Flash",         {"QWEN3-ASR-FLASH"}),
    ("Qwen3-ASR-1.7B",          {"QWEN3-ASR-1.7B"}),
    ("NVIDIA NeMo",             {"NVIDIA-NEMO"}),
    ("GPT-4o Transcribe",       {"GPT4O-TRANSCRIBE", "GPT-4O-TRANSCRIBE"}),
    ("Gemini 3.0 Flash",        {"GEMINI_3_0_FLASH", "GEMINI-3-FLASH-PREVIEW", "GEMINI", "GEMINI-3.0-FLASH"}),
    ("Whisper Large v3",        {"WHISPER", "WHISPER-LARGE-V3"}),
    ("Dolphin Small",           {"DOLPHIN_SMALL", "DOLPHIN-SMALL"}),
    ("Dolphin Base",            {"DOLPHIN_BASE", "DOLPHIN-BASE"}),
    ("FunASR-MLT-Nano",         {"FUN-ASR-MLT-NANO", "FUN-ASR-NANO", "FUNASR-MLT-NANO"}),
    ("FunASR-Realtime",         {"FUNASR-REALTIME", "FUN-ASR", "FUNASR_V1.5"}),
    ("Qwen3.5-Omni-Plus",       {"QWEN3.5-OMNI-PLUS", "QWEN3.5-OMNI-FLASH"}),
    ("BigASR",                  {"BIGASR"}),
    ("SeedASR",                 {"SEEDASR"}),
    ("Deepgram Nova 3",         {"DEEPGRAM_NOVA3", "DEEPGRAM-NOVA-3"}),
]
_INT2DISP = {i: d for d, internals in MODEL_ORDER for i in internals}
_DISP_ORDER = [d for d, _ in MODEL_ORDER]


# =========================
# Column order per sheet (country/language)
# =========================
EXCLUDE_COLS = {"EDU-EN", "EDU-CH"}

COLUMN_ORDER = {
    "Low-Resource-Languages": [
        "IRQ", "DZA", "ARE", "EGY", "MAR", "SAU", "SYR",
        "IDN", "MYS", "PHL", "PHL_EN", "PHL_noEN", "VNM", "THA",
        "JPN", "KOR",
    ],
    "CH-EN-Dialects": [
        "CHN-EN", "IDN-EN", "JPN-EN", "PHL-EN", "SCT-EN", "SGP-EN",
        "XIANG", "JIN", "GAN", "MIN", "YUE", "WU",
    ],
    "fleurs": [
        "EGY", "IDN", "MYS", "PHL", "VNM", "THA", "JPN", "KOR",
    ],
    "common-voice": [
        "AR", "IDN", "VNM", "THA", "JPN", "KOR",
    ],
    "Vertical-Domain-CH": [
        "AGR-CH", "AIT-CH", "ART-CH", "BIO-CH", "ECM-CH", "ENG-CH",
        "ENT-CH", "FIN-CH", "HUM-CH", "LAW-CH", "MED-CH", "MIL-CH",
    ],
    "Vertical-Domain-EN": [
        "AGR-EN", "AIT-EN", "ART-EN", "BIO-EN", "ECM-EN",
        "ENG-EN", "ENT-EN", "FIN-EN", "HUM-EN", "LAW-EN", "MED-EN", "MIL-EN",
    ],
    "Older-Children": [
        "CHILD-EN", "CHILD-CH", "OLD-EN", "OLD-CH",
    ],
}

# Fixed order: Low-Resource -> Dialects -> Vertical-Domain-CH -> Vertical-Domain-EN -> Older-Children
SHEET_ORDER = [
    "Low-Resource-Languages",
    "CH-EN-Dialects",
    "Vertical-Domain-CH",
    "Vertical-Domain-EN",
    "Older-Children",
]

# Hotword sheets: same column list as Vertical-Domain-{CH,EN}
COLUMN_ORDER["Hotword-CH"] = COLUMN_ORDER["Vertical-Domain-CH"]
COLUMN_ORDER["Hotword-EN"] = COLUMN_ORDER["Vertical-Domain-EN"]

# Hotword model display mapping (internal dir name -> display name).
# Directory naming differs slightly from results_Vertical-Domain (BIGASR has no suffix, SEEDASR, etc.)
HOTWORD_INT2DISP = {
    "AZURE": "Azure",
    "CHIRP3": "Chirp 3",
    "CHIRP-3": "Chirp 3",
    "ELEVENLABS_SCRIBE_V2": "ElevenLabs Scribe v2",
    "ELEVENLABS-SCRIBE-V2": "ElevenLabs Scribe v2",
    "OMNIASR_LLM_3B": "Meta OmniASR 3B",
    "META-OMNIASR-3B": "Meta OmniASR 3B",
    "QWEN3-ASR-FLASH": "Qwen3-ASR-Flash",
    "QWEN3-ASR-1.7B": "Qwen3-ASR-1.7B",
    "NVIDIA-NEMO": "NVIDIA NeMo",
    "GPT4O-TRANSCRIBE": "GPT-4o Transcribe",
    "GPT-4O-TRANSCRIBE": "GPT-4o Transcribe",
    "GEMINI": "Gemini 3.0 Flash",
    "GEMINI_3_0_FLASH": "Gemini 3.0 Flash",
    "GEMINI-3-FLASH-PREVIEW": "Gemini 3.0 Flash",
    "GEMINI-3.0-FLASH": "Gemini 3.0 Flash",
    "WHISPER": "Whisper Large v3",
    "WHISPER-LARGE-V3": "Whisper Large v3",
    "DOLPHIN_SMALL": "Dolphin Small",
    "DOLPHIN-SMALL": "Dolphin Small",
    "DOLPHIN_BASE": "Dolphin Base",
    "DOLPHIN-BASE": "Dolphin Base",
    "FUN-ASR-MLT-NANO": "FunASR-MLT-Nano",
    "FUN-ASR-NANO": "FunASR-MLT-Nano",
    "FUNASR-MLT-NANO": "FunASR-MLT-Nano",
    "FUN-ASR": "FunASR-Realtime",
    "FUNASR-REALTIME": "FunASR-Realtime",
    "QWEN3.5-OMNI-PLUS": "Qwen3.5-Omni-Plus",
    "QWEN3.5-OMNI-FLASH": "Qwen3.5-Omni-Plus",
    "BIGASR": "BigASR",
    "SEEDASR": "SeedASR",
    "DEEPGRAM_NOVA3": "Deepgram Nova 3",
    "DEEPGRAM-NOVA-3": "Deepgram Nova 3",
}

# Module -> results_<dir>
SOURCE_MAP = {
    "Vertical-Domain-CH": "Vertical-Domain",
    "Vertical-Domain-EN": "Vertical-Domain",
}

# Module -> data/text/<dir>/ref  (for computing valid ref duration)
TEXT_SOURCE_MAP = {
    "Low-Resource-Languages": "Low-Resource-Languages",
    "CH-EN-Dialects": "CH-EN-Dialects",
    "fleurs": "fleurs",
    "common-voice": "common-voice",
    "Vertical-Domain-CH": "Vertical-Domain",
    "Vertical-Domain-EN": "Vertical-Domain",
    "Older-Children": "Older-Children",
    "Hotword-CH": "Vertical-Domain",
    "Hotword-EN": "Vertical-Domain",
}


# =========================
# Utilities
# =========================
def normalize_model_name(model: str) -> str:
    m = re.match(r"^([A-Z]{3})_(.+)$", model)
    if m:
        rest = m.group(2)
        if rest in _INT2DISP:
            return rest
    return model


def pick_err_file(m_dir: str, suffix=""):
    """Pick the error file. suffix="" for original, suffix="-mindur" for filtered."""
    if not os.path.isdir(m_dir):
        return None
    names = [n for n in os.listdir(m_dir) if os.path.isfile(os.path.join(m_dir, n))]
    if not names:
        return None
    if suffix:
        # Look for files matching the suffix pattern
        err = [n for n in names if n.lower().startswith("err") and suffix in n]
    else:
        # Original: pick err files that do NOT have -mindur
        err = [n for n in names if n.lower().startswith("err") and "-mindur" not in n]
    if err:
        txts = [n for n in err if n.lower().endswith(".txt")]
        return os.path.join(m_dir, sorted(txts or err)[0])
    if not suffix:
        txts = sorted(n for n in names if n.lower().endswith(".txt") and "-mindur" not in n)
        return os.path.join(m_dir, txts[0]) if txts else None
    return None


def extract_wer(path: str):
    try:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("%WER"):
                    m = re.search(r"%WER\s*=\s*([0-9]+(?:\.[0-9]+)?)", line)
                    if m:
                        return float(m.group(1))
    except Exception:
        return None
    return None


def read_segment_check(m_dir: str):
    """Return (ref_n, matched_n) or None if missing."""
    p = os.path.join(m_dir, "segment_check.txt")
    if not os.path.isfile(p):
        return None
    vals = {}
    try:
        with open(p, "r", encoding="utf8") as f:
            for line in f:
                if "=" in line:
                    k, v = line.strip().split("=", 1)
                    vals[k.strip()] = v.strip()
        ref_n = int(vals.get("ref_segments", -1))
        matched_n = int(vals.get("matched_segments", -2))
        if ref_n < 0 or matched_n < 0:
            return None
        return ref_n, matched_n
    except Exception:
        return None


def load_ref_durations(base_dir: str, module: str, countries: list):
    """
    Read data/text/<source>/ref/<country>.json (flattened ref after valid filtering)
    and accumulate (end-start), returning {country: hours}. Returns None for missing files.
    """
    src = TEXT_SOURCE_MAP.get(module)
    out = {}
    if not src:
        return out
    ref_root = os.path.join(base_dir, "data", "text", src, "ref")
    for country in countries:
        path = os.path.join(ref_root, f"{country}.json")
        if not os.path.isfile(path):
            out[country] = None
            continue
        try:
            with open(path, "r", encoding="utf8") as f:
                data = json.load(f)
        except Exception:
            out[country] = None
            continue
        total_sec = 0.0
        if isinstance(data, list):
            for seg in data:
                try:
                    s = float(seg.get("start", 0))
                    e = float(seg.get("end", 0))
                    if e > s:
                        total_sec += (e - s)
                except Exception:
                    continue
        elif isinstance(data, dict):
            # Compatible with possible {audio: {segments:[...]}} structure
            for v in data.values():
                segs = v.get("segments", []) if isinstance(v, dict) else []
                for seg in segs:
                    if seg.get("status") == "invalid":
                        continue
                    try:
                        s = float(seg.get("start", 0))
                        e = float(seg.get("end", 0))
                        if e > s:
                            total_sec += (e - s)
                    except Exception:
                        continue
        out[country] = total_sec / 3600.0
    return out


# =========================
# Scan a module, return {display_model: {country: (cell_value, is_matched_or_None)}}
# is_matched: True/False/None  (None = no data)
# =========================
def scan_module(results_root: str, countries: list):
    table = {disp: {} for disp in _DISP_ORDER}
    for country in countries:
        c_dir = os.path.join(results_root, country)
        if not os.path.isdir(c_dir):
            continue
        for model in os.listdir(c_dir):
            m_dir = os.path.join(c_dir, model)
            if not os.path.isdir(m_dir):
                continue
            model_clean = normalize_model_name(model)
            display = _INT2DISP.get(model_clean)
            if not display:
                continue
            seg = read_segment_check(m_dir)
            err_path = pick_err_file(m_dir)
            wer = extract_wer(err_path) if err_path else None
            if seg is None:
                if wer is not None:
                    table[display][country] = (wer, True, None, None)
                continue
            ref_n, matched_n = seg
            if ref_n == matched_n:
                if wer is not None:
                    table[display][country] = (wer, True, ref_n, matched_n)
                else:
                    table[display][country] = ("-", None, ref_n, matched_n)
            else:
                # Alignment failed: keep WER and segment count info
                table[display][country] = (wer, False, ref_n, matched_n)
    return table


def prune_empty(table: dict, cols: list, models: list = None):
    """
    Remove columns where all models have no data, and models where all columns
    have no data. Returns (pruned_cols, pruned_models).
    """
    if models is None:
        models = list(_DISP_ORDER)

    def _has_data(entry):
        if entry is None:
            return False
        value = entry[0] if isinstance(entry, tuple) else entry
        return value is not None and value != "-"

    # Prune columns
    pruned_cols = []
    for col in cols:
        if any(_has_data(table.get(m, {}).get(col)) for m in models):
            pruned_cols.append(col)

    # Prune models
    pruned_models = []
    for m in models:
        if any(_has_data(table.get(m, {}).get(col)) for col in pruned_cols):
            pruned_models.append(m)

    return pruned_cols, pruned_models


def scan_module_filtered(results_root: str, countries: list):
    """Scan filtered (duration > threshold) results from -mindur files."""
    table = {disp: {} for disp in _DISP_ORDER}
    for country in countries:
        c_dir = os.path.join(results_root, country)
        if not os.path.isdir(c_dir):
            continue
        for model in os.listdir(c_dir):
            m_dir = os.path.join(c_dir, model)
            if not os.path.isdir(m_dir):
                continue
            model_clean = normalize_model_name(model)
            display = _INT2DISP.get(model_clean)
            if not display:
                continue
            seg = read_segment_check(m_dir)
            err_path = pick_err_file(m_dir, suffix="-mindur")
            wer = extract_wer(err_path) if err_path else None
            if seg is None:
                if wer is not None:
                    table[display][country] = (wer, True, None, None)
                continue
            ref_n, matched_n = seg
            if ref_n == matched_n:
                if wer is not None:
                    table[display][country] = (wer, True, ref_n, matched_n)
                else:
                    table[display][country] = ("-", None, ref_n, matched_n)
            else:
                table[display][country] = (wer, False, ref_n, matched_n)
    return table


# =========================
# Write a module to worksheet
# =========================
def write_sheet(ws, table: dict, columns: list, durations: dict = None, mode: str = "default", start_row: int = 1, section_title: str = None, models: list = None):
    """
    mode:
      "default"  - original behavior: fully aligned shows WER, misaligned shows matched/ref in red
      "besteff"  - best effort: show WER if available (even misaligned, in red), "-" only if no data
      "counts"   - only show matched/ref, misaligned in red
    start_row: starting row number (1-based)
    section_title: if not None, write a title row at start_row, data starts from start_row+1
    models: model display order list (default _DISP_ORDER)
    """
    if models is None:
        models = list(_DISP_ORDER)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    dur_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    durations = durations or {}

    offset = start_row - 1  # row offset from original (0-based addition)

    # Optional section title row
    if section_title:
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        tc = ws.cell(row=start_row, column=1, value=section_title)
        tc.font = Font(bold=True, size=12)
        tc.fill = title_fill
        for ci in range(2, len(columns) + 2):
            ws.cell(row=start_row, column=ci).fill = title_fill
        offset += 1  # shift data down by 1

    # Row 1: duration (valid ref)
    dc = ws.cell(row=1 + offset, column=1, value="Duration (valid ref)")
    dc.font = bold
    dc.alignment = center
    dc.fill = dur_fill
    for ci, col in enumerate(columns, start=2):
        h = durations.get(col)
        cell = ws.cell(row=1 + offset, column=ci)
        if h is None:
            cell.value = "-"
        else:
            cell.value = f"{h:.2f}h"
        cell.font = bold
        cell.alignment = center
        cell.fill = dur_fill

    # Row 2: header (Model / country)
    ws.cell(row=2 + offset, column=1, value="Model").font = bold
    ws.cell(row=2 + offset, column=1).alignment = center
    for ci, col in enumerate(columns, start=2):
        c = ws.cell(row=2 + offset, column=ci, value=col)
        c.font = bold
        c.alignment = center

    # Data rows (starting from row 3)
    for ri, disp in enumerate(models, start=3):
        ws.cell(row=ri + offset, column=1, value=disp).font = bold
        ws.cell(row=ri + offset, column=1).alignment = center
        for ci, col in enumerate(columns, start=2):
            entry = table.get(disp, {}).get(col)
            cell = ws.cell(row=ri + offset, column=ci)
            cell.alignment = center
            if entry is None:
                cell.value = "-"
                continue
            value, matched, ref_n, matched_n = entry

            if mode == "counts":
                # Only show matched/ref
                if ref_n is not None and matched_n is not None:
                    cell.value = f"{matched_n}/{ref_n}"
                    if matched is False:
                        cell.fill = red_fill
                        cell.font = red_font
                elif value == "-":
                    cell.value = "-"
                else:
                    cell.value = "-"
            elif mode == "besteff":
                # Best effort: show WER if available, misaligned in red
                if matched is True and isinstance(value, float):
                    cell.value = round(value, 2)
                    cell.number_format = "0.00"
                elif matched is False:
                    if isinstance(value, float):
                        cell.value = round(value, 2)
                        cell.number_format = "0.00"
                        cell.fill = red_fill
                        cell.font = red_font
                    elif matched_n is not None and matched_n > 0:
                        cell.value = f"{matched_n}/{ref_n}"
                        cell.fill = red_fill
                        cell.font = red_font
                    else:
                        cell.value = "-"
                else:
                    cell.value = "-" if value == "-" else value
            else:
                # default: original behavior
                if matched is False:
                    cell.value = f"{matched_n}/{ref_n}" if ref_n is not None else "-"
                    cell.fill = red_fill
                    cell.font = red_font
                elif matched is True and isinstance(value, float):
                    cell.value = round(value, 2)
                    cell.number_format = "0.00"
                else:
                    cell.value = value

    # MIN row (only fully aligned numeric columns — default/counts; besteff includes all values)
    min_row_idx = len(models) + 3 + offset
    ws.cell(row=min_row_idx, column=1, value="MIN").font = bold
    ws.cell(row=min_row_idx, column=1).alignment = center
    for ci, col in enumerate(columns, start=2):
        vals = []
        for disp in models:
            entry = table.get(disp, {}).get(col)
            if not entry:
                continue
            value, matched, ref_n, matched_n = entry
            if mode == "besteff":
                if isinstance(value, float):
                    vals.append(value)
            elif mode != "counts":
                if matched is True and isinstance(value, float):
                    vals.append(value)
                vals.append(entry[0])
        cell = ws.cell(row=min_row_idx, column=ci)
        cell.alignment = center
        cell.font = bold
        if vals:
            cell.value = round(min(vals), 2)
            cell.number_format = "0.00"
        else:
            cell.value = "-"

    # AVG columns: sub-group AVGs for specific sheets, then total AVG
    avg_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    subgrp_fill = PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid")  # light purple for sub-group AVG

    # Define sub-group AVG columns for specific sheets
    SUB_GROUPS = {
        "Low-Resource-Languages": [
            ("Arabic AVG", ["IRQ", "DZA", "ARE", "EGY", "MAR", "SAU", "SYR"]),
            ("SEA AVG", ["IDN", "MYS", "PHL", "PHL_EN", "PHL_noEN", "VNM", "THA"]),
            ("EA AVG", ["JPN", "KOR"]),
        ],
        "CH-EN-Dialects": [
            ("Accent AVG", ["CHN-EN", "IDN-EN", "JPN-EN", "PHL-EN", "SCT-EN", "SGP-EN"]),
            ("Dialect AVG", ["XIANG", "JIN", "GAN", "MIN", "YUE", "WU"]),
        ],
    }

    # Determine the sheet name from section_title or ws.title
    sheet_name = ws.title if ws.title else ""
    sub_groups = SUB_GROUPS.get(sheet_name, [])

    # Write sub-group AVG columns first, then total AVG
    next_ci = len(columns) + 2
    sub_avg_cols = []  # list of (col_index, group_name, group_columns)

    for grp_name, grp_cols in sub_groups:
        ci = next_ci
        sub_avg_cols.append((ci, grp_name, grp_cols))
        # Duration row
        ws.cell(row=1 + offset, column=ci, value="").fill = dur_fill
        ws.cell(row=1 + offset, column=ci).alignment = center
        # Header
        hdr = ws.cell(row=2 + offset, column=ci, value=grp_name)
        hdr.font = bold
        hdr.alignment = center
        hdr.fill = subgrp_fill
        # Model rows
        for ri, disp in enumerate(models, start=3):
            vals = []
            for col in grp_cols:
                if col not in columns:
                    continue
                entry = table.get(disp, {}).get(col)
                if not entry:
                    continue
                value, matched, ref_n, matched_n = entry
                if mode == "besteff":
                    if isinstance(value, float):
                        vals.append(value)
                elif mode != "counts":
                    if matched is True and isinstance(value, float):
                        vals.append(value)
            cell = ws.cell(row=ri + offset, column=ci)
            cell.alignment = center
            cell.fill = subgrp_fill
            if vals:
                cell.value = round(sum(vals) / len(vals), 2)
                cell.number_format = "0.00"
                cell.font = bold
            else:
                cell.value = "-"
        # MIN row
        min_vals = []
        for disp in models:
            cv = ws.cell(row=models.index(disp) + 3 + offset, column=ci).value
            if isinstance(cv, (int, float)):
                min_vals.append(cv)
        min_cell = ws.cell(row=min_row_idx, column=ci)
        min_cell.alignment = center
        min_cell.font = bold
        min_cell.fill = subgrp_fill
        if min_vals:
            min_cell.value = round(min(min_vals), 2)
            min_cell.number_format = "0.00"
        else:
            min_cell.value = "-"
        next_ci += 1

    # Total AVG column
    avg_ci = next_ci
    # Duration row
    ws.cell(row=1 + offset, column=avg_ci, value="").fill = dur_fill
    ws.cell(row=1 + offset, column=avg_ci).alignment = center
    # Header
    hdr = ws.cell(row=2 + offset, column=avg_ci, value="AVG")
    hdr.font = bold
    hdr.alignment = center
    hdr.fill = avg_fill
    # Model rows
    for ri, disp in enumerate(models, start=3):
        vals = []
        for col in columns:
            entry = table.get(disp, {}).get(col)
            if not entry:
                continue
            value, matched, ref_n, matched_n = entry
            if mode == "besteff":
                if isinstance(value, float):
                    vals.append(value)
            elif mode == "counts":
                pass  # no avg for counts
            else:
                if matched is True and isinstance(value, float):
                    vals.append(value)
        cell = ws.cell(row=ri + offset, column=avg_ci)
        cell.alignment = center
        cell.fill = avg_fill
        if vals:
            cell.value = round(sum(vals) / len(vals), 2)
            cell.number_format = "0.00"
            cell.font = bold
        else:
            cell.value = "-"
    # MIN row AVG
    min_avgs = []
    for disp in models:
        cell_val = ws.cell(row=models.index(disp) + 3 + offset, column=avg_ci).value
        if isinstance(cell_val, (int, float)):
            min_avgs.append(cell_val)
    min_avg_cell = ws.cell(row=min_row_idx, column=avg_ci)
    min_avg_cell.alignment = center
    min_avg_cell.font = bold
    min_avg_cell.fill = avg_fill
    if min_avgs:
        min_avg_cell.value = round(min(min_avgs), 2)
        min_avg_cell.number_format = "0.00"
    else:
        min_avg_cell.value = "-"

    # Highlight best (red) and second-best (yellow) per column
    # Ties: all tied-best get red, yellow goes to next distinct value
    best_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    best_font = Font(color="FFFFFF", bold=True)
    second_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    second_font = Font(bold=True)

    # All data columns + sub-group AVGs + total AVG
    all_highlight_cols = list(range(2, avg_ci + 1))
    for ci in all_highlight_cols:
        col_vals = []
        for ri, disp in enumerate(models, start=3):
            cell = ws.cell(row=ri + offset, column=ci)
            if isinstance(cell.value, (int, float)):
                col_vals.append((cell.value, ri + offset))
        if not col_vals:
            continue
        col_vals.sort(key=lambda x: x[0])
        best_val = col_vals[0][0]
        # Find all tied-best
        best_rows = [row for val, row in col_vals if abs(val - best_val) < 0.005]
        for brow in best_rows:
            bc = ws.cell(row=brow, column=ci)
            if bc.fill != red_fill:
                bc.fill = best_fill
                bc.font = best_font
        # Find next distinct value after best
        second_val = None
        for val, row in col_vals:
            if abs(val - best_val) >= 0.005:
                second_val = val
                break
        if second_val is not None:
            second_rows = [row for val, row in col_vals if abs(val - second_val) < 0.005]
            for srow in second_rows:
                sc = ws.cell(row=srow, column=ci)
                if sc.fill != red_fill:
                    sc.fill = second_fill
                    sc.font = second_font

    # Column widths (only set on first call)
    if start_row == 1:
        ws.column_dimensions["A"].width = 26
        for ci in range(2, avg_ci + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

    # Return the next available row (after MIN row + 1 blank)
    return min_row_idx + 2


# =========================
# Hotword B-WER/B-CER scanner (for appending to main VD sheets)
# =========================
def scan_hotword_bwer(results_root: str, domains: list, wer_results_root: str = None):
    """
    Scan hotword results and return a table in scan_module format:
    {display_model: {domain: (b_wer_value, is_matched, total, matched_n)}}

    is_matched is determined by the normal WER segment_check when available,
    so B-WER cells are green/red-consistent with the main WER table.
    """
    # Pre-load normal WER segment checks for each (domain, model)
    normal_seg = {}  # (display, domain) -> (is_matched, ref_n, matched_n)
    if wer_results_root and os.path.isdir(wer_results_root):
        for domain in domains:
            d_dir = os.path.join(wer_results_root, domain)
            if not os.path.isdir(d_dir):
                continue
            for model in os.listdir(d_dir):
                m_dir = os.path.join(d_dir, model)
                if not os.path.isdir(m_dir):
                    continue
                model_clean = normalize_model_name(model)
                display = _INT2DISP.get(model_clean)
                if not display:
                    continue
                seg = read_segment_check(m_dir)
                if seg is not None:
                    ref_n, matched_n = seg
                    normal_seg[(display, domain)] = (ref_n == matched_n, ref_n, matched_n)

    table = {disp: {} for disp in _DISP_ORDER}
    for domain in domains:
        d_dir = os.path.join(results_root, domain)
        if not os.path.isdir(d_dir):
            continue
        for model in os.listdir(d_dir):
            m_dir = os.path.join(d_dir, model)
            if not os.path.isdir(m_dir):
                continue
            model_clean = normalize_model_name(model)
            display = HOTWORD_INT2DISP.get(model_clean)
            if not display:
                continue
            p = os.path.join(m_dir, "hotword_result.txt")
            if not os.path.isfile(p):
                continue
            parsed = _parse_hotword_result(p)
            if parsed is None:
                continue
            b_wer = parsed["metrics"].get("b_wer")
            if b_wer is None:
                continue
            total = parsed["matched"] + parsed["skipped"]
            # Use normal WER segment check for match status and counts
            seg_info = normal_seg.get((display, domain))
            if seg_info is not None:
                is_matched, ref_n, matched_n = seg_info
                table[display][domain] = (b_wer, is_matched, ref_n, matched_n)
            else:
                is_matched = parsed["skipped"] == 0
                table[display][domain] = (b_wer, is_matched, total, parsed["matched"])
    return table


# =========================
# Hotword result scanning and writing
# =========================
HOTWORD_METRICS = [
    ("WER (%)", "CER (%)", "wer"),
    ("U-WER (%)", "U-CER (%)", "u_wer"),
    ("B-WER (%)", "B-CER (%)", "b_wer"),
    ("Recall (%)", "Recall (%)", "recall"),
]


def _parse_hotword_result(path: str):
    out = {}
    try:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                line = line.strip()
                if "=" in line:
                    k, v = line.split("=", 1)
                    out[k.strip()] = v.strip()
    except Exception:
        return None
    try:
        matched = int(out.get("matched_segments", "0"))
        skipped = int(out.get("skipped_segments", "0"))
    except ValueError:
        return None
    metrics = {}
    for key in ("wer", "u_wer", "b_wer", "recall"):
        v = out.get(key)
        if v is None:
            continue
        try:
            metrics[key] = float(v)
        except ValueError:
            continue
    return {"matched": matched, "skipped": skipped, "metrics": metrics}


def scan_hotword(results_root: str, domains: list):
    """Return {display_model: {domain: {metric: (value, is_matched)}}}."""
    # Display models by hotword directory name
    table = {}
    display_order = []
    for domain in domains:
        d_dir = os.path.join(results_root, domain)
        if not os.path.isdir(d_dir):
            continue
        for model in sorted(os.listdir(d_dir)):
            m_dir = os.path.join(d_dir, model)
            if not os.path.isdir(m_dir):
                continue
            model_clean = normalize_model_name(model)
            display = HOTWORD_INT2DISP.get(model_clean)
            if not display:
                continue
            if display not in table:
                table[display] = {}
                if display not in display_order:
                    display_order.append(display)
            p = os.path.join(m_dir, "hotword_result.txt")
            if not os.path.isfile(p):
                continue
            parsed = _parse_hotword_result(p)
            if parsed is None:
                continue
            total = parsed["matched"] + parsed["skipped"]
            is_matched = parsed["skipped"] == 0
            cell_map = {}
            for _, _, key in HOTWORD_METRICS:
                v = parsed["metrics"].get(key)
                if v is None:
                    cell_map[key] = None
                elif is_matched:
                    cell_map[key] = (v, True, total, parsed["matched"])
                else:
                    cell_map[key] = (v, False, total, parsed["matched"])
            table[display][domain] = cell_map

    # Sort by _DISP_ORDER
    ordered = [d for d in _DISP_ORDER if d in table]
    ordered += [d for d in display_order if d not in _DISP_ORDER]
    return table, ordered


def write_hotword_sheet(ws, table: dict, ordered_models: list, columns: list, durations: dict = None, mode: str = "default", is_chinese: bool = False):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    dur_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    durations = durations or {}

    # Top row: duration (valid ref)
    dc = ws.cell(row=1, column=1, value="Duration (valid ref)")
    dc.font = bold
    dc.alignment = center
    dc.fill = dur_fill
    for ci, col in enumerate(columns, start=2):
        h = durations.get(col)
        cell = ws.cell(row=1, column=ci)
        cell.value = "-" if h is None else f"{h:.2f}h"
        cell.font = bold
        cell.alignment = center
        cell.fill = dur_fill

    row = 2
    for en_label, ch_label, metric_key in HOTWORD_METRICS:
        metric_label = ch_label if is_chinese else en_label
        # Section title
        ws.cell(row=row, column=1, value=metric_label).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        for ci in range(2, len(columns) + 2):
            ws.cell(row=row, column=ci).fill = section_fill
        row += 1

        # Header row
        ws.cell(row=row, column=1, value="Model").font = bold
        ws.cell(row=row, column=1).alignment = center
        for ci, col in enumerate(columns, start=2):
            c = ws.cell(row=row, column=ci, value=col)
            c.font = bold
            c.alignment = center
        row += 1

        # Data rows
        for disp in ordered_models:
            ws.cell(row=row, column=1, value=disp).font = bold
            ws.cell(row=row, column=1).alignment = center
            for ci, col in enumerate(columns, start=2):
                cell_map = table.get(disp, {}).get(col)
                cell = ws.cell(row=row, column=ci)
                cell.alignment = center
                if not cell_map:
                    cell.value = "-"
                    continue
                entry = cell_map.get(metric_key)
                if entry is None:
                    cell.value = "-"
                    continue
                value, is_matched, total, matched_n = entry
                if mode == "counts":
                    cell.value = f"{matched_n}/{total}"
                    if not is_matched:
                        cell.fill = red_fill
                        cell.font = red_font
                elif mode == "besteff":
                    if isinstance(value, float):
                        cell.value = round(value, 2)
                        cell.number_format = "0.00"
                        if not is_matched:
                            cell.fill = red_fill
                            cell.font = red_font
                    else:
                        cell.value = "-"
                else:
                    if not is_matched:
                        cell.value = f"{matched_n}/{total}"
                        cell.fill = red_fill
                        cell.font = red_font
                    elif isinstance(value, float):
                        cell.value = round(value, 2)
                        cell.number_format = "0.00"
                    else:
                        cell.value = value
            row += 1

        # MIN row (only for besteff/default with WER metric)
        if mode != "counts" and metric_key == "wer":
            ws.cell(row=row, column=1, value="MIN").font = bold
            ws.cell(row=row, column=1).alignment = center
            for ci, col in enumerate(columns, start=2):
                vals = []
                for disp in ordered_models:
                    cell_map = table.get(disp, {}).get(col)
                    if not cell_map:
                        continue
                    entry = cell_map.get(metric_key)
                    if entry is None:
                        continue
                    value, is_matched, total, matched_n = entry
                    if mode == "besteff":
                        if isinstance(value, float):
                            vals.append(value)
                    else:
                        if is_matched and isinstance(value, float):
                            vals.append(value)
                cell = ws.cell(row=row, column=ci)
                cell.alignment = center
                cell.font = bold
                if vals:
                    cell.value = round(min(vals), 2)
                    cell.number_format = "0.00"
                else:
                    cell.value = "-"
            row += 1

        # blank separator row
        row += 1

    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# =========================
# main
# =========================
def main(base_dir: str):
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(base_dir, "data", "excel_output")
    os.makedirs(out_dir, exist_ok=True)

    # Generate both Excel files
    for mode, suffix in [("besteff", "_besteff"), ("counts", "_counts")]:
        out_path = os.path.join(out_dir, f"all_results{suffix}_{timestamp}.xlsx")
        wb = Workbook()
        wb.remove(wb.active)

        for module in SHEET_ORDER:
            source = SOURCE_MAP.get(module, module)
            # Use only mindur (>0.5s) results
            results_root = os.path.join(base_dir, "data", f"results_{source}_mindur")
            if not os.path.isdir(results_root):
                # Fallback to original if mindur not available
                results_root = os.path.join(base_dir, "data", f"results_{source}")
                if not os.path.isdir(results_root):
                    continue
            cols = [c for c in COLUMN_ORDER[module] if c not in EXCLUDE_COLS]
            table = scan_module(results_root, cols)

            # Prune empty columns and models
            cols, active_models = prune_empty(table, cols)

            matched_cnt = unmatched_cnt = empty_cnt = 0
            for disp in active_models:
                for col in cols:
                    entry = table.get(disp, {}).get(col)
                    if entry is None:
                        empty_cnt += 1
                    elif entry[1] is True:
                        matched_cnt += 1
                    elif entry[1] is False:
                        unmatched_cnt += 1
                    else:
                        empty_cnt += 1

            ws = wb.create_sheet(title=module[:31])
            durations = load_ref_durations(base_dir, module, cols)
            next_row = write_sheet(ws, table, cols, durations, mode, models=active_models)

            # Append B-WER/B-CER section for Vertical-Domain sheets
            if module in ("Vertical-Domain-CH", "Vertical-Domain-EN"):
                hotword_mindur_root = os.path.join(base_dir, "data", "results_hotword_mindur")
                if os.path.isdir(hotword_mindur_root):
                    is_ch = module.endswith("-CH")
                    bwer_label = "B-CER (%)" if is_ch else "B-WER (%)"
                    bwer_table = scan_hotword_bwer(hotword_mindur_root, cols, wer_results_root=results_root)
                    next_row = write_sheet(ws, bwer_table, cols, durations, mode,
                                start_row=next_row,
                                section_title=bwer_label,
                                models=active_models)

            if mode == "besteff":
                print(
                    f"✅ {module}: matched={matched_cnt}  "
                    f"unmatched(red)={unmatched_cnt}  empty={empty_cnt}  cols={len(cols)}"
                )

        wb.save(out_path)
        print(f"\n📄 Written -> {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Merge per-module results into one Excel")
    parser.add_argument(
        "--base_dir", type=str,
        default=os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        help="Project root (default: auto-detect from script location)",
    )
    args = parser.parse_args()
    main(args.base_dir)
