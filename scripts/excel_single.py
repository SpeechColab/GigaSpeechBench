#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generate WER/CER results Excel from evaluation outputs.

Scans results_root/{COUNTRY}/{MODEL}/ directories, extracts WER/CER values,
and produces a single Excel file with model-vs-country matrix.
"""

import os
import re
import json
import pandas as pd
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Model whitelist
COVERAGE_MODELS = {
    "AZURE","BIGASR_V400","CHIRP3","DOLPHIN_SMALL","DOLPHIN_BASE",
    "ELEVENLABS_SCRIBE_V2","FUN-ASR-MLT-NANO","GEMINI_3_0_FLASH","GPT4O-TRANSCRIBE",
    "OMNIASR_LLM_3B","QWEN3-ASR-FLASH","STT_AR_FASTCONFORMER_HYBRID_LARGE_PCD_V1.0.NEMO",
    "WHISPER","STT_KR_CONFORMER_TRANSDUCER_LARGE","PARAKEET-TDT_CTC-0.6B-JA",
    "NVIDIA-NEMO","WHISPER-LARGE-V3","GEMINI","GEMINI_3_0_FLASH","SEEDASR_2.0",
    "GEMINI-3-FLASH-PREVIEW","QWEN3-ASR-1.7B","FUN-ASR-NANO","SEEDASR2","SEEDASR",
    "FUN-ASR","QWEN3.5-OMNI-FLASH","FUNASR_V1.5"
}

# Model display order: (display_name, {internal_names...})
MODEL_ORDER = [
    ("Azure",                   {"AZURE"}),
    ("Chirp3",                  {"CHIRP3"}),
    ("elevenlabs_scribe_v2",    {"ELEVENLABS_SCRIBE_V2"}),
    ("meta(omniASR_LLM_3B)",    {"OMNIASR_LLM_3B"}),
    ("qwen3-asr-flash",         {"QWEN3-ASR-FLASH"}),
    ("qwen3-asr",               {"QWEN3-ASR-1.7B"}),
    ("nvidia-nemo",             {"NVIDIA-NEMO"}),
    ("gpt4o-transcribe",        {"GPT4O-TRANSCRIBE"}),
    ("gemini 3.0 flash",        {"GEMINI_3_0_FLASH", "GEMINI-3-FLASH-PREVIEW", "GEMINI"}),
    ("whisper",                  {"WHISPER", "WHISPER-LARGE-V3"}),
    ("dolphin_small",           {"DOLPHIN_SMALL"}),
    ("dolphin_base",            {"DOLPHIN_BASE"}),
    ("fun-asr-mlt-nano",        {"FUN-ASR-MLT-NANO", "FUN-ASR-NANO"}),
    ("funasr1.5",               {"FUN-ASR", "FUNASR_V1.5"}),
    ("qwen3.5-omni-flash",      {"QWEN3.5-OMNI-FLASH"}),
    ("seedasr-1-BIGASR_V400",   {"BIGASR_V400", "SEEDASR"}),
    ("SEEDASR_2.0",             {"SEEDASR_2.0", "SEEDASR2"}),
]

def _build_internal_to_display():
    """internal name -> display name"""
    m = {}
    for display, internals in MODEL_ORDER:
        for i in internals:
            m[i] = display
    return m

_INT2DISP = _build_internal_to_display()
_DISP_ORDER = [d for d, _ in MODEL_ORDER]

# =========================
# Utilities
# =========================
def extract_value(path: str):
    """Extract %WER/CER value from error file."""
    try:
        with open(path, "r", encoding="utf8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("%WER"):
                    m = re.search(r"%WER\s*=\s*([0-9]+(?:\.[0-9]+)?)", line)
                    if m:
                        return float(m.group(1))
        return "-"
    except Exception:
        return "-"

def normalize_model_name(model: str):
    # Strip 3-letter country prefix (e.g. ARE_, DZA_) but NOT model names like FUNASR_
    # Only strip if the prefix looks like a country code followed by a known model
    m = re.match(r"^([A-Z]{3})_(.+)$", model)
    if m:
        prefix, rest = m.group(1), m.group(2)
        # Check if rest (after stripping prefix) is a known model
        if rest in COVERAGE_MODELS:
            return rest
    return model

def pick_err_file(m_dir: str):
    if not os.path.isdir(m_dir):
        return None
    names = [n for n in os.listdir(m_dir) if os.path.isfile(os.path.join(m_dir, n))]
    if not names:
        return None
    err = [n for n in names if n.lower().startswith("err")]
    if err:
        txts = [n for n in err if n.lower().endswith(".txt")]
        return os.path.join(m_dir, sorted(txts or err)[0])
    txts = sorted(n for n in names if n.lower().endswith(".txt"))
    return os.path.join(m_dir, txts[0]) if txts else None

# =========================
# Scan results
# =========================
def _is_fully_matched(m_dir: str) -> bool:
    """Check segment_check.txt: return True only if matched == ref."""
    seg_path = os.path.join(m_dir, "segment_check.txt")
    if not os.path.isfile(seg_path):
        return False
    vals = {}
    try:
        with open(seg_path, "r", encoding="utf8") as f:
            for line in f:
                if "=" in line:
                    k, v = line.strip().split("=", 1)
                    vals[k] = v
        ref_n = int(vals.get("ref_segments", -1))
        matched_n = int(vals.get("matched_segments", -2))
        return ref_n == matched_n
    except Exception:
        return False

def scan_results(results_root: str, excel_countries, matched_only: bool = False):
    table = {}
    for country in excel_countries:
        c_dir = os.path.join(results_root, country)
        if not os.path.isdir(c_dir):
            continue
        for model in os.listdir(c_dir):
            m_dir = os.path.join(c_dir, model)
            if not os.path.isdir(m_dir):
                continue
            model_clean = normalize_model_name(model)
            if model_clean not in COVERAGE_MODELS:
                continue
            if matched_only and not _is_fully_matched(m_dir):
                continue
            err_path = pick_err_file(m_dir)
            if not err_path:
                continue
            val = extract_value(err_path)
            # Map to display name; merge same-display models into one row
            display = _INT2DISP.get(model_clean, model_clean)
            table.setdefault(display, {})[country] = val
    df = pd.DataFrame.from_dict(table, orient="index")
    df = df.reindex(columns=excel_countries)
    # Strict MODEL_ORDER sort; all models appear even if no data ("-")
    df = df.reindex(_DISP_ORDER)
    return df.fillna("-")

# =========================
# Ref segment counts
# =========================
def load_ref_counts(ref_root: str):
    """Load ref segment counts: total, valid (non-empty text), and duration (hours)."""
    ref_count = {}
    ref_valid_count = {}
    ref_duration = {}
    for fn in os.listdir(ref_root):
        if not fn.endswith(".json"):
            continue
        country = fn[:-5]
        path = os.path.join(ref_root, fn)
        try:
            with open(path, "r", encoding="utf8") as f:
                data = json.load(f)
            if isinstance(data, list):
                ref_count[country] = len(data)
                ref_valid_count[country] = sum(1 for d in data if d.get("text", "").strip())
                total_dur = sum(
                    float(d.get("end", 0)) - float(d.get("start", 0))
                    for d in data if d.get("text", "").strip()
                )
                ref_duration[country] = round(total_dur / 3600, 2)
        except Exception:
            continue
    return ref_count, ref_valid_count, ref_duration


def scan_matched_counts(results_root: str, excel_countries):
    """Build matched segment count table from segment_check.txt files."""
    table = {}
    for country in excel_countries:
        c_dir = os.path.join(results_root, country)
        if not os.path.isdir(c_dir):
            continue
        for model in os.listdir(c_dir):
            m_dir = os.path.join(c_dir, model)
            if not os.path.isdir(m_dir):
                continue
            model_clean = normalize_model_name(model)
            if model_clean not in COVERAGE_MODELS:
                continue
            seg_path = os.path.join(m_dir, "segment_check.txt")
            if not os.path.isfile(seg_path):
                continue
            vals = {}
            try:
                with open(seg_path, "r", encoding="utf8") as f:
                    for line in f:
                        if "=" in line:
                            k, v = line.strip().split("=", 1)
                            vals[k] = v
                matched_n = int(vals.get("matched_segments", 0))
            except Exception:
                continue
            display = _INT2DISP.get(model_clean, model_clean)
            table.setdefault(display, {})[country] = matched_n
    df = pd.DataFrame.from_dict(table, orient="index")
    df = df.reindex(columns=excel_countries)
    df = df.reindex(_DISP_ORDER)
    return df.fillna(0).astype(int)


# =========================
# Main
# =========================
def main(results_root: str, ref_root: str, excel_countries, skip_existing: bool = False, matched_only: bool = False):
    print(f"Scanning results: {results_root}")
    if matched_only:
        print("Mode: matched_only (only fully aligned results)")
    ref_count, ref_valid_count, ref_duration = load_ref_counts(ref_root)
    df = scan_results(results_root, excel_countries, matched_only=matched_only)

    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ========== WER/CER Excel ==========
    out_xlsx = os.path.join(results_root, "results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "WER_CER"

    # Header row
    ws.cell(row=1, column=1, value="")
    for ci, c in enumerate(excel_countries, start=2):
        ws.cell(row=1, column=ci, value=c)

    # Row 2: Ref Valid Segments
    ws.cell(row=2, column=1, value="Ref Valid Segments")
    for ci, c in enumerate(excel_countries, start=2):
        ws.cell(row=2, column=ci, value=ref_valid_count.get(c, 0))

    # Row 3: Duration (hours)
    ws.cell(row=3, column=1, value="Duration (hours)")
    for ci, c in enumerate(excel_countries, start=2):
        ws.cell(row=3, column=ci, value=ref_duration.get(c, 0))

    # Model rows (row 4+)
    model_start_row = 4
    for ri, model_name in enumerate(df.index, start=model_start_row):
        ws.cell(row=ri, column=1, value=model_name)
        for ci, c in enumerate(excel_countries, start=2):
            val = df.at[model_name, c] if c in df.columns else "-"
            ws.cell(row=ri, column=ci, value=val)

    # Min row at the bottom
    min_row_idx = model_start_row + len(df)
    ws.cell(row=min_row_idx, column=1, value="MIN")
    for ci, c in enumerate(excel_countries, start=2):
        col_vals = []
        for ri in range(model_start_row, min_row_idx):
            v = ws.cell(row=ri, column=ci).value
            if isinstance(v, (int, float)):
                col_vals.append((v, ri))
        if col_vals:
            min_val, min_ri = min(col_vals, key=lambda x: x[0])
            ws.cell(row=min_row_idx, column=ci, value=min_val)
            # Highlight the min cell in yellow
            ws.cell(row=min_ri, column=ci).fill = YELLOW_FILL

    wb.save(out_xlsx)
    print(f"Wrote WER/CER Excel: {out_xlsx}")

    # ========== Count Excel ==========
    out_count_xlsx = os.path.join(results_root, "results_count.xlsx")
    df_count = scan_matched_counts(results_root, excel_countries)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Matched_Count"

    # Header row
    ws2.cell(row=1, column=1, value="")
    for ci, c in enumerate(excel_countries, start=2):
        ws2.cell(row=1, column=ci, value=c)

    # Row 2: Ref Valid Segments
    ws2.cell(row=2, column=1, value="Ref Valid Segments")
    for ci, c in enumerate(excel_countries, start=2):
        ws2.cell(row=2, column=ci, value=ref_valid_count.get(c, 0))

    # Row 3: Duration (hours)
    ws2.cell(row=3, column=1, value="Duration (hours)")
    for ci, c in enumerate(excel_countries, start=2):
        ws2.cell(row=3, column=ci, value=ref_duration.get(c, 0))

    # Model rows (row 4+)
    for ri, model_name in enumerate(df_count.index, start=4):
        ws2.cell(row=ri, column=1, value=model_name)
        for ci, c in enumerate(excel_countries, start=2):
            val = int(df_count.at[model_name, c]) if c in df_count.columns else 0
            cell = ws2.cell(row=ri, column=ci, value=val)
            # Red if matched < ref valid
            ref_v = ref_valid_count.get(c, 0)
            if ref_v > 0 and val < ref_v:
                cell.fill = RED_FILL

    wb2.save(out_count_xlsx)
    print(f"Wrote Count Excel: {out_count_xlsx}")

    print("\nRef segment counts:")
    for c, n in sorted(ref_count.items()):
        valid = ref_valid_count.get(c, 0)
        print(f"  {c}: total={n}, valid={valid}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Strict non-batch Excel generator")
    parser.add_argument("--results_root", type=str, required=True)
    parser.add_argument("--ref_root", type=str, required=True)
    parser.add_argument("--excel_countries", type=str, nargs="+", required=True,
                        help="Countries/regions for Excel, e.g. AGR-CH AIT-CH ...")
    parser.add_argument("--skip_existing", type=int, choices=[0, 1], default=0,
                        help="1: skip existing Excel; 0: overwrite")
    parser.add_argument("--matched_only", type=int, choices=[0, 1], default=0,
                        help="1: only include fully matched (ref==matched) results; 0: include all")
    args = parser.parse_args()
    main(args.results_root, args.ref_root, args.excel_countries, bool(args.skip_existing), bool(args.matched_only))