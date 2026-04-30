#!/usr/bin/env python3
"""
Speech Translation evaluation pipeline.
Scans all st_results* directories, evaluates each model, and outputs
two Excel files: one for EN target, one for ZH target.
Each Excel has one sheet per metric. Each row = one model, columns = languages + OVERALL.

Supported metrics:
  - BLEU, chrF, TER  (sacrebleu, always available, fast)
  - COMET             (unbabel-comet, needs model download ~1.5GB, CPU ok)
  - BLEURT            (bleurt-pytorch, needs model download ~1.2GB, CPU ok)

Usage:
  # All lightweight metrics (default)
  python3 eval_st_pipeline.py

  # Add COMET (first run downloads model)
  python3 eval_st_pipeline.py --comet

  # Add BLEURT
  python3 eval_st_pipeline.py --bleurt

  # All metrics
  python3 eval_st_pipeline.py --comet --bleurt

  # Only specific metrics
  python3 eval_st_pipeline.py --metrics BLEU COMET
"""

import argparse
import gc
import glob
import json
import os
import sys

import jieba
import sacrebleu
import torch
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --------------- config ---------------

ALL_LANGS = ["ARE", "DZA", "EGY", "IDN", "IRQ", "MAR", "MYS", "PHL", "SAU", "THA", "VNM"]

MODEL_NAME_MAP = {
    "st_results": "Gemini-2.0-Flash",
    "st_results_qwen3lt": "Qwen3-LiveTranslate-Flash",
}

BASE_METRICS = ["BLEU", "chrF", "TER"]
NEURAL_METRICS = ["COMET", "BLEURT"]
ALL_METRICS = BASE_METRICS + NEURAL_METRICS

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")

DEFAULT_COMET_MODEL = "Unbabel/wmt22-comet-da"
DEFAULT_BLEURT_MODEL = "lucadiliello/BLEURT-20"


# --------------- tokenizer ---------------

def tokenize_zh(text: str) -> str:
    return " ".join(jieba.cut(text))


# --------------- data loading ---------------

def load_file(path: str):
    with open(path, encoding="utf-8") as f:
        return json.load(f) or []


def collect_data(results_dir: str, lang: str, target: str):
    """Load refs, hyps, sources for one (lang, target)."""
    path = os.path.join(results_dir, f"{lang}_{target}.json")
    if not os.path.isfile(path):
        return None
    data = load_file(path)
    if not data:
        return None
    return (
        [d["ref"] for d in data],
        [d["hyp"] for d in data],
        [d.get("original", "") for d in data],
    )


# --------------- metric computers ---------------

def compute_base_metrics(refs, hyps, is_zh=False):
    if is_zh:
        refs_tok = [tokenize_zh(r) for r in refs]
        hyps_tok = [tokenize_zh(h) for h in hyps]
        bleu = sacrebleu.corpus_bleu(hyps_tok, [refs_tok], tokenize="none")
    else:
        bleu = sacrebleu.corpus_bleu(hyps, [refs])
    chrf = sacrebleu.corpus_chrf(hyps, [refs])
    ter = sacrebleu.corpus_ter(hyps, [refs])
    return {
        "BLEU": round(bleu.score, 2),
        "chrF": round(chrf.score, 2),
        "TER": round(ter.score, 2),
    }


class CometScorer:
    def __init__(self, model_name=DEFAULT_COMET_MODEL, device="cpu"):
        from comet import download_model, load_from_checkpoint
        print(f"  Loading COMET: {model_name} (device={device})")
        model_path = download_model(model_name)
        self.model = load_from_checkpoint(model_path)
        if device == "cuda" and torch.cuda.is_available():
            self.model = self.model.to("cuda")
        self.gpus = 1 if (device == "cuda" and torch.cuda.is_available()) else 0

    def score(self, srcs, hyps, refs):
        data = [{"src": s, "mt": h, "ref": r} for s, h, r in zip(srcs, hyps, refs)]
        output = self.model.predict(data, batch_size=8, gpus=self.gpus)
        return round(float(output.system_score), 4)

    def cleanup(self):
        del self.model
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()


class BleurtScorer:
    def __init__(self, model_name=DEFAULT_BLEURT_MODEL, device="cpu"):
        from bleurt_pytorch import BleurtForSequenceClassification, BleurtTokenizer
        print(f"  Loading BLEURT: {model_name} (device={device})")
        self.tokenizer = BleurtTokenizer.from_pretrained(model_name)
        self.model = BleurtForSequenceClassification.from_pretrained(model_name)
        self.device = device if (device == "cuda" and torch.cuda.is_available()) else "cpu"
        self.model = self.model.to(self.device).eval()

    def score(self, hyps, refs):
        all_scores = []
        for i in range(0, len(refs), 32):
            br, bh = refs[i:i+32], hyps[i:i+32]
            with torch.no_grad():
                inputs = self.tokenizer(br, bh, padding="longest", truncation=True,
                                        max_length=512, return_tensors="pt")
                inputs = {k: v.to(self.device) for k, v in inputs.items()}
                scores = self.model(**inputs).logits.flatten().tolist()
                all_scores.extend(scores)
        return round(float(np.mean(all_scores)), 4)

    def cleanup(self):
        del self.model, self.tokenizer
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()


# --------------- discovery ---------------

def discover_models(data_dir: str):
    models = []
    for d in sorted(glob.glob(os.path.join(data_dir, "st_results*"))):
        if not os.path.isdir(d):
            continue
        dirname = os.path.basename(d)
        model_name = MODEL_NAME_MAP.get(dirname, dirname)
        jsons = glob.glob(os.path.join(d, "*_en.json")) + glob.glob(os.path.join(d, "*_zh.json"))
        if jsons:
            models.append((d, model_name))
    return models


# --------------- Excel ---------------

def build_excel(models_data, target, output_path, langs, metrics):
    wb = Workbook()
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    overall_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for si, metric in enumerate(metrics):
        ws = wb.active if si == 0 else wb.create_sheet()
        ws.title = metric

        headers = ["Model"] + langs + ["OVERALL"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for ri, (model_name, lang_metrics, overall) in enumerate(models_data, 2):
            ws.cell(row=ri, column=1, value=model_name).font = bold
            ws.cell(row=ri, column=1).alignment = center

            for ci, lang in enumerate(langs, 2):
                cell = ws.cell(row=ri, column=ci)
                m = lang_metrics.get(lang)
                if m and metric in m:
                    cell.value = m[metric]
                    cell.number_format = "0.0000" if metric in NEURAL_METRICS else "0.00"
                else:
                    cell.value = "-"
                cell.alignment = center

            # OVERALL column
            oc = len(langs) + 2
            cell = ws.cell(row=ri, column=oc)
            if overall and metric in overall:
                cell.value = overall[metric]
                cell.number_format = "0.0000" if metric in NEURAL_METRICS else "0.00"
                cell.font = bold
                cell.fill = overall_fill
            else:
                cell.value = "-"
            cell.alignment = center

        ws.column_dimensions["A"].width = 28
        for ci in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 12

    wb.save(output_path)


# --------------- main ---------------

def main():
    parser = argparse.ArgumentParser(
        description="ST evaluation pipeline: all models -> two Excel files (EN/ZH)")
    parser.add_argument("--data_dir", type=str, default=DATA_DIR)
    parser.add_argument("--output_dir", type=str, default=None)
    parser.add_argument("--comet", action="store_true", help="Enable COMET metric")
    parser.add_argument("--bleurt", action="store_true", help="Enable BLEURT metric")
    parser.add_argument("--metrics", nargs="+", default=None, choices=ALL_METRICS,
                        help="Select specific metrics (overrides --comet/--bleurt)")
    parser.add_argument("--device", type=str, default="cpu", choices=["cpu", "cuda"])
    args = parser.parse_args()

    data_dir = os.path.abspath(args.data_dir)
    output_dir = args.output_dir or data_dir

    # Determine active metrics
    if args.metrics:
        active_metrics = args.metrics
    else:
        active_metrics = list(BASE_METRICS)
        if args.comet:
            active_metrics.append("COMET")
        if args.bleurt:
            active_metrics.append("BLEURT")

    use_comet = "COMET" in active_metrics
    use_bleurt = "BLEURT" in active_metrics

    models = discover_models(data_dir)
    if not models:
        print(f"No st_results* directories found in {data_dir}")
        sys.exit(1)

    print(f"Models:  {len(models)}")
    for d, name in models:
        n = len(glob.glob(os.path.join(d, "*.json")))
        print(f"  {name:<30} ({n} files)")
    print(f"Metrics: {', '.join(active_metrics)}")
    print(f"Device:  {args.device}")

    # Load neural scorers once
    comet_scorer = CometScorer(device=args.device) if use_comet else None
    bleurt_scorer = BleurtScorer(device=args.device) if use_bleurt else None

    for target in ["en", "zh"]:
        is_zh = (target == "zh")
        label = "English" if target == "en" else "Chinese"
        print(f"\n{'='*60}")
        print(f"Target: {label}")
        print(f"{'='*60}")

        models_data = []
        for results_dir, model_name in models:
            print(f"\n  [{model_name}]")
            lang_metrics = {}

            for lang in ALL_LANGS:
                triplet = collect_data(results_dir, lang, target)
                if not triplet:
                    continue
                refs, hyps, srcs = triplet
                m = {}

                if any(k in active_metrics for k in BASE_METRICS):
                    m.update(compute_base_metrics(refs, hyps, is_zh=is_zh))

                if use_comet and comet_scorer and any(srcs):
                    try:
                        m["COMET"] = comet_scorer.score(srcs, hyps, refs)
                    except Exception as e:
                        print(f"    COMET err {lang}: {e}")

                if use_bleurt and bleurt_scorer:
                    try:
                        m["BLEURT"] = bleurt_scorer.score(hyps, refs)
                    except Exception as e:
                        print(f"    BLEURT err {lang}: {e}")

                m["n_segments"] = len(refs)
                lang_metrics[lang] = m
                vals = ", ".join(f"{k}={m[k]}" for k in active_metrics if k in m)
                print(f"    {lang}: {len(refs)} segs  {vals}")

            # Overall
            all_refs, all_hyps, all_srcs = [], [], []
            for lang in ALL_LANGS:
                triplet = collect_data(results_dir, lang, target)
                if triplet:
                    r, h, s = triplet
                    all_refs.extend(r); all_hyps.extend(h); all_srcs.extend(s)

            overall = {}
            if all_refs:
                if any(k in active_metrics for k in BASE_METRICS):
                    overall.update(compute_base_metrics(all_refs, all_hyps, is_zh=is_zh))
                if use_comet and comet_scorer and any(all_srcs):
                    try:
                        overall["COMET"] = comet_scorer.score(all_srcs, all_hyps, all_refs)
                    except Exception as e:
                        print(f"    COMET overall err: {e}")
                if use_bleurt and bleurt_scorer:
                    try:
                        overall["BLEURT"] = bleurt_scorer.score(all_hyps, all_refs)
                    except Exception as e:
                        print(f"    BLEURT overall err: {e}")
                overall["n_segments"] = len(all_refs)

            models_data.append((model_name, lang_metrics, overall))
            if overall:
                vals = ", ".join(f"{k}={overall.get(k,'N/A')}" for k in active_metrics)
                print(f"    OVERALL: {overall.get('n_segments',0)} segs  {vals}")

        out_path = os.path.join(output_dir, f"st_eval_{target}.xlsx")
        build_excel(models_data, target, out_path, ALL_LANGS, active_metrics)
        print(f"\n  Excel: {out_path}")

    if comet_scorer:
        comet_scorer.cleanup()
    if bleurt_scorer:
        bleurt_scorer.cleanup()
    print(f"\nDone.")


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Speech Translation evaluation pipeline.
Scans all st_results* directories, evaluates each model with sacreBLEU, chrF++, TER,
and outputs two Excel files: one for EN target, one for ZH target.
Each row = one model, columns = per-language metrics + OVERALL.

Usage:
  python3 eval_st_pipeline.py
  python3 eval_st_pipeline.py --output_dir ./data/st_results
"""

import argparse
import glob
import json
import os
import sys

import jieba
import sacrebleu
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# All possible languages (superset across models)
ALL_LANGS = ["ARE", "DZA", "EGY", "IDN", "IRQ", "MAR", "MYS", "PHL", "SAU", "THA", "VNM"]

# Map result directory name -> display model name
MODEL_NAME_MAP = {
    "st_results": "Gemini-2.0-Flash",
    "st_results_qwen3lt": "Qwen3-LiveTranslate-Flash",
}

METRICS = ["BLEU", "chrF", "TER"]

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "data")


def tokenize_zh(text: str) -> str:
    return " ".join(jieba.cut(text))


def evaluate_file(path: str, is_zh: bool = False):
    """Evaluate a single result file. Returns dict with metrics."""
    try:
        data = json.load(open(path, encoding="utf-8"))
    except Exception:
        return None
    if not data:
        return None

    refs = [item["ref"] for item in data]
    hyps = [item["hyp"] for item in data]

    if is_zh:
        refs_tok = [tokenize_zh(r) for r in refs]
        hyps_tok = [tokenize_zh(h) for h in hyps]
        bleu = sacrebleu.corpus_bleu(hyps_tok, [refs_tok], tokenize="none")
    else:
        bleu = sacrebleu.corpus_bleu(hyps, [refs])

    chrf = sacrebleu.corpus_chrf(hyps, [refs])
    ter = sacrebleu.corpus_ter(hyps, [refs])

    return {
        "n_segments": len(data),
        "BLEU": round(bleu.score, 2),
        "chrF": round(chrf.score, 2),
        "TER": round(ter.score, 2),
    }


def evaluate_overall(results_dir: str, target: str, langs: list):
    """Compute overall metrics across all languages for a target."""
    is_zh = (target == "zh")
    all_refs, all_hyps = [], []
    for lang in langs:
        path = os.path.join(results_dir, f"{lang}_{target}.json")
        if not os.path.isfile(path):
            continue
        data = json.load(open(path, encoding="utf-8"))
        all_refs.extend([d["ref"] for d in data])
        all_hyps.extend([d["hyp"] for d in data])

    if not all_refs:
        return None

    if is_zh:
        bleu = sacrebleu.corpus_bleu(
            [tokenize_zh(h) for h in all_hyps],
            [[tokenize_zh(r) for r in all_refs]],
            tokenize="none",
        )
    else:
        bleu = sacrebleu.corpus_bleu(all_hyps, [all_refs])

    chrf = sacrebleu.corpus_chrf(all_hyps, [all_refs])
    ter = sacrebleu.corpus_ter(all_hyps, [all_refs])

    return {
        "n_segments": len(all_refs),
        "BLEU": round(bleu.score, 2),
        "chrF": round(chrf.score, 2),
        "TER": round(ter.score, 2),
    }


def discover_models(data_dir: str):
    """Discover all st_results* directories and return list of (dir_path, model_name)."""
    models = []
    for d in sorted(glob.glob(os.path.join(data_dir, "st_results*"))):
        if not os.path.isdir(d):
            continue
        dirname = os.path.basename(d)
        model_name = MODEL_NAME_MAP.get(dirname, dirname)
        # Check it has at least one JSON result
        jsons = glob.glob(os.path.join(d, "*_en.json")) + glob.glob(os.path.join(d, "*_zh.json"))
        if jsons:
            models.append((d, model_name))
    return models


def build_excel(models_data, target: str, output_path: str, langs: list):
    """
    Build Excel file for one target language.
    3 sheets: BLEU, chrF, TER. Each sheet: rows=models, columns=languages + OVERALL.
    models_data: list of (model_name, {lang: metrics_dict}, overall_metrics_dict)
    """
    wb = Workbook()

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    overall_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for si, metric in enumerate(METRICS):
        if si == 0:
            ws = wb.active
            ws.title = metric
        else:
            ws = wb.create_sheet(title=metric)

        # Headers: Model | ARE | DZA | ... | VNM | OVERALL
        headers = ["Model"] + langs + ["OVERALL"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data rows
        for ri, (model_name, lang_metrics, overall) in enumerate(models_data, 2):
            ws.cell(row=ri, column=1, value=model_name).font = bold
            ws.cell(row=ri, column=1).alignment = center

            for ci, lang in enumerate(langs, 2):
                cell = ws.cell(row=ri, column=ci)
                m = lang_metrics.get(lang)
                if m:
                    cell.value = m[metric]
                    cell.number_format = "0.00"
                else:
                    cell.value = "-"
                cell.alignment = center

            # OVERALL column
            oc = len(langs) + 2
            cell = ws.cell(row=ri, column=oc)
            if overall:
                cell.value = overall[metric]
                cell.number_format = "0.00"
                cell.font = bold
                cell.fill = overall_fill
            else:
                cell.value = "-"
            cell.alignment = center

        # Column widths
        ws.column_dimensions["A"].width = 28
        for ci in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="ST evaluation pipeline: all models → two Excel files (EN/ZH)")
    parser.add_argument("--data_dir", type=str, default=DATA_DIR,
                        help="Parent dir containing st_results* folders")
    parser.add_argument("--output_dir", type=str, default=None,
                        help="Output directory for Excel files (default: data_dir)")
    args = parser.parse_args()

    data_dir = os.path.abspath(args.data_dir)
    output_dir = args.output_dir or data_dir

    models = discover_models(data_dir)
    if not models:
        print(f"No st_results* directories found in {data_dir}")
        sys.exit(1)

    print(f"Found {len(models)} model(s):")
    for d, name in models:
        n_files = len(glob.glob(os.path.join(d, "*.json")))
        print(f"  {name:<30} ({n_files} files in {os.path.basename(d)}/)")

    # Evaluate all models × targets
    for target in ["en", "zh"]:
        is_zh = (target == "zh")
        label = "English" if target == "en" else "Chinese"
        print(f"\n{'='*60}")
        print(f"Evaluating → {label}")
        print(f"{'='*60}")

        models_data = []
        for results_dir, model_name in models:
            lang_metrics = {}
            for lang in ALL_LANGS:
                path = os.path.join(results_dir, f"{lang}_{target}.json")
                if os.path.isfile(path):
                    m = evaluate_file(path, is_zh=is_zh)
                    if m:
                        lang_metrics[lang] = m

            overall = evaluate_overall(results_dir, target, ALL_LANGS)
            models_data.append((model_name, lang_metrics, overall))

            # Print summary
            n_langs = len(lang_metrics)
            if overall:
                print(f"  {model_name:<30} {n_langs} langs, "
                      f"BLEU={overall['BLEU']:.2f}, chrF={overall['chrF']:.2f}, "
                      f"TER={overall['TER']:.2f} ({overall['n_segments']} segs)")
            else:
                print(f"  {model_name:<30} no data for {target}")

        # Build Excel
        out_path = os.path.join(output_dir, f"st_eval_{target}.xlsx")
        build_excel(models_data, target, out_path, ALL_LANGS)
        print(f"  Excel: {out_path}")

    print(f"\nDone.")


if __name__ == "__main__":
    main()
